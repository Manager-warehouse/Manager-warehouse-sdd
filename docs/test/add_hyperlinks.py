# -*- coding: utf-8 -*-
"""
Script to add clickable sheet hyperlinks in Test Cases and Test Statistics sheets
"""

import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(r'd:\swp\Manager-warehouse-sdd')
EXCEL_OUT = ROOT / 'docs' / 'test' / 'test_final.xlsx'

def add_hyperlinks():
    wb = openpyxl.load_workbook(str(EXCEL_OUT))

    tahoma_regular = Font(name='Tahoma', size=10, bold=False)
    tahoma_link = Font(name='Tahoma', size=10, bold=False, color='0000FF', underline='single')
    tahoma_bold_link = Font(name='Tahoma', size=10, bold=True, color='0000FF', underline='single')

    align_left = Alignment(horizontal='left', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')
    align_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )

    # 1. Update Test Cases Sheet
    ws_tc = wb['Test Cases']
    for r in range(9, 19):
        func_name = ws_tc.cell(row=r, column=3).value
        sheet_code = ws_tc.cell(row=r, column=4).value

        # Clean formula or text if previously set
        if isinstance(func_name, str) and 'HYPERLINK' in func_name:
            # Extract plain text from HYPERLINK formula
            parts = func_name.split('"')
            if len(parts) >= 4:
                func_name = parts[-2]

        if isinstance(sheet_code, str) and 'HYPERLINK' in sheet_code:
            parts = sheet_code.split('"')
            if len(parts) >= 4:
                sheet_code = parts[-2]

        if sheet_code and sheet_code in wb.sheetnames:
            # Set Function Name as clickable hyperlink
            c_func = ws_tc.cell(row=r, column=3)
            c_func.value = f'=HYPERLINK("#\'{sheet_code}\'!A1", "{func_name}")'
            c_func.font = tahoma_link
            c_func.alignment = align_left
            c_func.border = thin_border

            # Set Sheet Name as clickable hyperlink
            c_sheet = ws_tc.cell(row=r, column=4)
            c_sheet.value = f'=HYPERLINK("#\'{sheet_code}\'!A1", "{sheet_code}")'
            c_sheet.font = tahoma_link
            c_sheet.alignment = align_center
            c_sheet.border = thin_border

    # 2. Update Test Statistics Sheet
    ws_stat = wb['Test Statistics']
    for r in range(11, 21):
        sname_cell = ws_stat.cell(row=r, column=3)
        # Row index to module code
        sname = wb.sheetnames[r - 8]  # index 3 for row 11 (AUTH-001)
        if sname in wb.sheetnames:
            sname_cell.value = f'=HYPERLINK("#\'{sname}\'!A1", {sname}!B1)'
            sname_cell.font = tahoma_link
            sname_cell.alignment = align_left
            sname_cell.border = thin_border

    # 3. Add a "Return to Index / Test Cases" link on row 1 of each Workflow sheet!
    for sname in wb.sheetnames[3:]:
        ws_mod = wb[sname]
        # Set cell O1 (Column 15) to Return link
        ret_cell = ws_mod.cell(row=1, column=15)
        ret_cell.value = '=HYPERLINK("#\'Test Cases\'!A1", "Back to Test Cases List")'
        ret_cell.font = Font(name='Tahoma', size=9, bold=True, color='0000FF', underline='single')
        ret_cell.alignment = Alignment(horizontal='right', vertical='center')

    wb.save(str(EXCEL_OUT))
    print(f"Successfully added sheet hyperlinks to {EXCEL_OUT}")

if __name__ == '__main__':
    add_hyperlinks()
