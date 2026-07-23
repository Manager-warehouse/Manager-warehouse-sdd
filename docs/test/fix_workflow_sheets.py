# -*- coding: utf-8 -*-
"""
Script to fix Workflow sheets to match Template3_System Test.xlsx EXACTLY:
1. Metadata block (Rows 2 to 8): Workflow name in B2, Requirement in B3, Number of TCs in B4, Round 1-3 in B6:E8
2. Row 10 TABLE HEADER: Test Case ID | Test Case Description | Test Case Procedure | Expected Results | Pre-conditions | Round 1 | Test date | Tester | Round 2 | Test date | Tester | Round 3 | Test date | Tester | Note
   (Preserving exact Olive Green fill and Bold White font from template!)
3. Scenario Rows: Preserving exact Light Aqua/Cyan fill from template across columns 1-15
4. Data Rows starting from row 12+
"""

import os
import re
import sys
import copy
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink

ROOT = Path(r'd:\swp\Manager-warehouse-sdd')
TEMPLATE = ROOT / 'docs' / 'test' / 'Template3_System Test.xlsx'
SUREFIRE = ROOT / 'backend' / 'target' / 'surefire-reports'
BE_TEST_DIR = ROOT / 'backend' / 'src' / 'test' / 'java' / 'com' / 'wms'
FE_TEST_DIRS = [
    ROOT / 'frontend' / 'src',
    ROOT / 'frontend' / 'test_frontend',
]
EXCEL_OUT = ROOT / 'docs' / 'test' / 'test_final.xlsx'
RUN_DATE = datetime.now().strftime('%Y-%m-%d')
TESTER_AUTO = 'CI/Automated'

PROJECT_NAME = 'Warehouse Management System (WMS)'
PROJECT_CODE = 'WMS-SPRINT1'

MODULES = [
    {
        'id': '001', 'name': 'Security, Auth & RBAC', 'code': 'AUTH-001',
        'req': 'Kiểm tra xác thực người dùng, phân quyền theo role, kiểm soát truy cập kho, cấu hình hệ thống và audit log',
        'be_keywords': ['Auth', 'Security', 'User', 'AuditLog', 'SystemConfig', 'WarehouseIsolation', 'RBAC'],
        'fe_keywords': ['rbac', 'config', 'admin'],
        'scenarios': [
            'Đăng nhập hệ thống',
            'Làm mới token (Refresh Token)',
            'Đổi mật khẩu',
            'Quản lý người dùng (User CRUD)',
            'Phân quyền theo Role (RBAC)',
            'Kiểm soát truy cập kho',
            'Cấu hình hệ thống (System Config)',
            'Audit Log',
        ],
    },
    {
        'id': '002', 'name': 'Master Data Management', 'code': 'MDM-002',
        'req': 'Kiểm tra quản lý dữ liệu danh mục: sản phẩm, kho, vị trí kho, nhà cung cấp, đại lý, xe, tài xế',
        'be_keywords': ['Product', 'Warehouse', 'WarehouseLocation', 'Supplier', 'Dealer', 'Vehicle', 'Driver'],
        'fe_keywords': ['ProductManagement', 'masterData'],
        'scenarios': ['Quản lý sản phẩm', 'Quản lý kho', 'Quản lý vị trí kho (Zone & Bin)',
                      'Quản lý nhà cung cấp', 'Quản lý đại lý', 'Quản lý xe', 'Quản lý tài xế'],
    },
    {
        'id': '003', 'name': 'Inbound Receipt & QC', 'code': 'RCV-003',
        'req': 'Kiểm tra toàn bộ quy trình nhập kho: tạo phiếu, phê duyệt, QC, putaway, quarantine, RTV',
        'be_keywords': ['Receipt', 'ReceiptQc', 'ReceiptPutaway', 'Quarantine', 'InboundReceipt',
                        'ReceiptValidation', 'ReceiptBackend', 'ReceiptRtv'],
        'fe_keywords': ['ReceiptList', 'QuarantineWorkspace', 'inbound'],
        'scenarios': ['Tạo phiếu nhập kho', 'Phê duyệt phiếu nhập', 'QC Inbound',
                      'Putaway hàng vào Bin', 'Xử lý hàng Quarantine', 'Trả hàng NCC (RTV)',
                      'Tìm kiếm & lọc phiếu nhập'],
    },
    {
        'id': '004', 'name': 'Outbound Delivery & POD', 'code': 'OUT-004',
        'req': 'Kiểm tra quy trình xuất kho: Delivery Order, QC outbound, chuyến giao hàng, POD',
        'be_keywords': ['DeliveryOrder', 'Trip', 'DriverDelivery', 'TripService', 'TripController'],
        'fe_keywords': ['DeliveryOrders', 'QCOutbound', 'DriverTrip', 'outbound'],
        'scenarios': ['Tạo Delivery Order', 'QC Outbound', 'Tạo chuyến giao hàng (Trip)',
                      'Xuất phát chuyến (Depart)', 'Xác nhận giao hàng (POD)',
                      'Hủy chuyến hàng', 'Hoàn thành chuyến (Complete)'],
    },
    {
        'id': '005', 'name': 'Inter-Warehouse Transfer', 'code': 'TRF-005',
        'req': 'Kiểm tra quy trình điều chuyển hàng giữa các kho: yêu cầu, phê duyệt, QC, In-Transit, nhận hàng',
        'be_keywords': ['TransferRequest', 'Transfer', 'InterWarehouseTransfer'],
        'fe_keywords': ['TransferRequest', 'inter-warehouse-transfer', 'validation'],
        'scenarios': ['Tạo yêu cầu điều chuyển', 'Submit & Approve yêu cầu', 'Chuyển thành Transfer',
                      'QC Outbound tại kho nguồn', 'Bàn giao hàng cho tài xế',
                      'Nhận hàng tại kho đích', 'QC nhận hàng tại kho đích',
                      'Xử lý hàng trả về nguồn', 'Chênh lệch số lượng'],
    },
    {
        'id': '006', 'name': 'Stocktake & Adjustment', 'code': 'STK-006',
        'req': 'Kiểm tra quy trình kiểm kê kho, xử lý chênh lệch và điều chỉnh tồn kho',
        'be_keywords': ['StockTake', 'Adjustment'],
        'fe_keywords': ['StocktakeList', 'stocktake'],
        'scenarios': ['Tạo phiếu kiểm kê', 'Nhập số lượng thực đếm', 'Submit phiếu kiểm kê',
                      'Phê duyệt chênh lệch (Variance)', 'Điều chỉnh tồn kho (Adjustment)',
                      'Lọc & tìm kiếm phiếu kiểm kê'],
    },
    {
        'id': '007', 'name': 'Pricing & COGS', 'code': 'PRC-007',
        'req': 'Kiểm tra quản lý giá bán, lịch sử giá và tính giá vốn hàng bán (COGS theo FIFO)',
        'be_keywords': ['PriceHistory', 'Price'],
        'fe_keywords': ['pricing'],
        'scenarios': ['Tạo bảng giá (Price History)', 'Phê duyệt giá', 'Xem lịch sử giá', 'Tính COGS'],
    },
    {
        'id': '008', 'name': 'Finance, Billing & Closing', 'code': 'FIN-008',
        'req': 'Kiểm tra quy trình tài chính: hóa đơn, thanh toán, công nợ đại lý, đóng kỳ kế toán',
        'be_keywords': ['Invoice', 'Payment', 'AccountingPeriod', 'AutoInvoice', 'SupplierInvoice',
                        'SupplierPayment', 'BillingNotification'],
        'fe_keywords': ['PeriodClosing', 'SupplierInvoices', 'DealerDebtInvoice', 'PaymentReceipts', 'finance'],
        'scenarios': ['Tạo hóa đơn tự động (Auto Invoice)', 'Hóa đơn NCC (Supplier Invoice)',
                      'Thanh toán NCC', 'Quản lý công nợ đại lý', 'Đóng kỳ kế toán (Period Closing)',
                      'Báo cáo công nợ (Credit Aging)'],
    },
    {
        'id': '009', 'name': 'Returns, Scrap & Disposal', 'code': 'RET-009',
        'req': 'Kiểm tra quy trình trả hàng khách, cấp credit note và xử lý hàng tiêu hủy',
        'be_keywords': ['Returns', 'Disposal', 'ReceiptRtv', 'ReceiptServiceReturn'],
        'fe_keywords': ['returns'],
        'scenarios': ['Tạo phiếu trả hàng khách', 'QC hàng trả', 'Cấp Credit Note', 'Xử lý hàng tiêu hủy'],
    },
    {
        'id': '010', 'name': 'Reports, Dashboard & Alerts', 'code': 'RPT-010',
        'req': 'Kiểm tra báo cáo, dashboard CEO và cảnh báo tồn kho thấp',
        'be_keywords': ['Report', 'StockAlert', 'ReportService'],
        'fe_keywords': ['report', 'CreditAgingReport'],
        'scenarios': ['CEO Dashboard', 'Báo cáo định giá tồn kho', 'Cảnh báo tồn kho thấp', 'Báo cáo công nợ'],
    },
]

def parse_surefire(surefire_dir: Path) -> dict:
    results = {}
    if not surefire_dir.exists():
        return results
    today = datetime.now().strftime('%Y-%m-%d')
    for xf in surefire_dir.glob('TEST-*.xml'):
        try:
            tree = ET.parse(xf)
            root = tree.getroot()
            suite = root.get('name', '')
            ts = root.get('timestamp', today)
            if ts and 'T' in ts:
                ts = ts.split('T')[0]
            for tc in root.findall('testcase'):
                method = tc.get('name', '')
                cls = tc.get('classname', suite).split('.')[-1]
                if tc.find('failure') is not None or tc.find('error') is not None:
                    st = 'Failed'
                elif tc.find('skipped') is not None:
                    st = 'N/A'
                else:
                    st = 'Passed'
                results.setdefault(cls, {})[method] = {'status': st, 'date': ts}
        except Exception:
            continue
    return results

def snake_to_words(name: str) -> str:
    parts = name.split('_')
    words = []
    for p in parts:
        words.append(re.sub(r'([A-Z])', r' \1', p).strip().lower())
    return ' '.join(words)

def make_procedure(method: str) -> str:
    w = snake_to_words(method).lower()
    if any(x in w for x in ['login', 'auth', 'signin']):
        return '1. Mở trình duyệt, truy cập trang đăng nhập\n2. Nhập thông tin tài khoản\n3. Nhấn Đăng nhập'
    elif any(x in w for x in ['create', 'add']):
        return '1. Đăng nhập với role phù hợp\n2. Điều hướng đến chức năng\n3. Nhấn Tạo mới\n4. Điền thông tin\n5. Nhấn Lưu / Submit'
    elif any(x in w for x in ['update', 'edit']):
        return '1. Đăng nhập với role phù hợp\n2. Mở bản ghi cần cập nhật\n3. Nhấn Chỉnh sửa\n4. Thay đổi thông tin\n5. Nhấn Lưu'
    elif any(x in w for x in ['delete', 'deactivate', 'disable']):
        return '1. Đăng nhập với role phù hợp\n2. Tìm bản ghi cần xử lý\n3. Nhấn Vô hiệu hóa / Xóa\n4. Xác nhận thao tác'
    elif any(x in w for x in ['approve', 'phê duyệt', 'duyệt']):
        return '1. Đăng nhập với role có quyền duyệt\n2. Mở danh sách chờ duyệt\n3. Chọn phiếu\n4. Nhấn Phê duyệt'
    elif any(x in w for x in ['reject', 'cancel', 'hủy', 'từ chối']):
        return '1. Đăng nhập với role phù hợp\n2. Mở phiếu cần xử lý\n3. Nhấn Từ chối / Hủy\n4. Nhập lý do'
    elif any(x in w for x in ['list', 'search', 'filter', 'getall', 'tìm']):
        return '1. Đăng nhập hệ thống\n2. Điều hướng đến danh sách\n3. Áp dụng bộ lọc\n4. Quan sát kết quả'
    else:
        return f'1. Đăng nhập hệ thống\n2. Thực hiện: {w[:60]}\n3. Quan sát kết quả'

def make_expected(method: str) -> str:
    w = method.lower()
    if any(x in w for x in ['success', 'valid', 'returns200', 'returns201', 'returns204']):
        return 'Hệ thống xử lý thành công. Thông báo success hiển thị. Dữ liệu được lưu đúng.'
    elif any(x in w for x in ['rejects', 'fails', 'returns400', 'throws', 'invalid', 'error']):
        return 'Hệ thống hiển thị thông báo lỗi. Không có dữ liệu bị thay đổi. HTTP 4xx trả về.'
    elif any(x in w for x in ['returns401', 'unauthenticated', 'unauthorized']):
        return 'Hệ thống từ chối truy cập. HTTP 401 Unauthorized.'
    elif any(x in w for x in ['returns403', 'forbidden']):
        return 'Hệ thống từ chối quyền. HTTP 403 Forbidden.'
    elif any(x in w for x in ['notfound', 'returns404']):
        return 'Hệ thống trả về không tìm thấy. HTTP 404.'
    else:
        return 'Hệ thống phản hồi đúng nghiệp vụ. Dữ liệu hiển thị chính xác.'

def extract_precond(java_content: str) -> str:
    match = re.search(
        r'@BeforeEach[\s\S]*?void\s+\w+\s*\(\s*\)\s*\{([\s\S]*?)(?=\n\s*@|\Z)',
        java_content)
    conds = []
    if match:
        body = match.group(1)
        if re.search(r'new\s+User|mockUser|adminUser|managerUser', body):
            conds.append('Tài khoản người dùng đã được tạo')
        if re.search(r'new\s+Warehouse|mockWarehouse|warehouse\s*=', body):
            conds.append('Kho hàng đã được khởi tạo')
        if re.search(r'new\s+Product|mockProduct|product\s*=', body):
            conds.append('Sản phẩm đã được đăng ký')
        if re.search(r'JWT|token|jwtToken|authHeader', body):
            conds.append('Token xác thực hợp lệ đã được cấp phát')
        if re.search(r'new\s+Receipt|receipt\s*=', body):
            conds.append('Phiếu nhập kho đã tồn tại')
        if re.search(r'new\s+Trip|trip\s*=', body):
            conds.append('Chuyến hàng đã được khởi tạo')
    if not conds:
        conds = ['Hệ thống đang hoạt động bình thường', 'Người dùng đã đăng nhập với role phù hợp']
    return '; '.join(conds)

def get_tcs_for_module(module: dict, surefire: dict) -> list:
    rows = []
    for root, _, files in os.walk(BE_TEST_DIR):
        for f in files:
            if not f.endswith('.java'):
                continue
            cls = f.replace('.java', '')
            if not any(kw.lower() in cls.lower() for kw in module['be_keywords']):
                continue
            path = Path(root) / f
            content = path.read_text(encoding='utf-8', errors='ignore')
            precond = extract_precond(content)
            methods = re.findall(r'@Test[\s\S]{0,300}?void\s+(\w+)\s*\(', content)
            cls_results = surefire.get(cls, {})
            for method in methods:
                res = cls_results.get(method, {})
                st = res.get('status', 'Pending')
                dt = res.get('date', '')
                tester = TESTER_AUTO if st in ('Passed', 'Failed') else ''
                scenario_idx = (len(rows) // 5) % len(module['scenarios'])
                scenario = module['scenarios'][scenario_idx]
                rows.append({
                    'scenario': scenario,
                    'id': f"{module['code']}-{len(rows)+1:03d}",
                    'desc': snake_to_words(method).title(),
                    'proc': make_procedure(method),
                    'exp': make_expected(method),
                    'pre': precond,
                    'r1': st,
                    'dt': dt,
                    'tester': tester,
                    'note': f'Backend: {cls}.java',
                })

    pattern = re.compile(r"""(?:it|test)\s*\(['"](.*?)['"]\s*,""")
    for base in FE_TEST_DIRS:
        if not base.exists():
            continue
        for root, _, files in os.walk(base):
            for f in files:
                if '.test.' not in f:
                    continue
                if not any(kw.lower() in f.lower() for kw in module['fe_keywords']):
                    continue
                path = Path(root) / f
                content = path.read_text(encoding='utf-8', errors='ignore')
                for m in pattern.findall(content):
                    scenario_idx = (len(rows) // 5) % len(module['scenarios'])
                    scenario = module['scenarios'][scenario_idx]
                    rows.append({
                        'scenario': scenario,
                        'id': f"{module['code']}-FE{len(rows)+1:03d}",
                        'desc': m,
                        'proc': make_procedure(m),
                        'exp': make_expected(m),
                        'pre': 'Trình duyệt đã mở ứng dụng WMS; Người dùng đã đăng nhập',
                        'r1': 'Pending',
                        'dt': '',
                        'tester': '',
                        'note': f'Frontend: {f}',
                    })
    return rows

def copy_cell_style(src, dst):
    if src.has_style:
        if src.font:
            dst.font = copy.copy(src.font)
        if src.fill:
            dst.fill = copy.copy(src.fill)
        if src.border:
            dst.border = copy.copy(src.border)
        if src.alignment:
            dst.alignment = copy.copy(src.alignment)
        if src.number_format:
            dst.number_format = src.number_format

def fix_workflow_sheets():
    print("Loading Template3_System Test.xlsx...")
    wb_template = openpyxl.load_workbook(str(TEMPLATE))
    ws_tmpl = wb_template['Workflow Name1']

    wb_out = openpyxl.load_workbook(str(EXCEL_OUT))

    surefire = parse_surefire(SUREFIRE)

    tahoma_regular = Font(name='Tahoma', size=10, bold=False, color='000000')
    tahoma_bold = Font(name='Tahoma', size=10, bold=True, color='000000')
    data_align_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
    data_align_center = Alignment(horizontal='center', vertical='center')

    status_fills = {
        'Passed': PatternFill('solid', fgColor='C6EFCE'),
        'Failed': PatternFill('solid', fgColor='FFC7CE'),
        'Pending': PatternFill('solid', fgColor='FFEB9C'),
        'N/A': PatternFill('solid', fgColor='D9D9D9'),
    }

    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )

    for mod in MODULES:
        code = mod['code']
        ws = wb_out[code]

        # 1. Metadata Block (Rows 2 to 8) - Copy exact style from Template
        ws['A2'] = 'Workflow'
        copy_cell_style(ws_tmpl.cell(2, 1), ws['A2'])
        ws['B2'] = mod['name']
        copy_cell_style(ws_tmpl.cell(2, 2), ws['B2'])

        ws['A3'] = 'Test requirement'
        copy_cell_style(ws_tmpl.cell(3, 1), ws['A3'])
        ws['B3'] = mod['req']
        copy_cell_style(ws_tmpl.cell(3, 2), ws['B3'])

        ws['A4'] = 'Number of TCs'
        copy_cell_style(ws_tmpl.cell(4, 1), ws['A4'])
        ws['B4'] = '=COUNTA(A12:A5000)'
        copy_cell_style(ws_tmpl.cell(4, 2), ws['B4'])

        # Row 5: Testing Round Headers
        ws['A5'] = 'Testing Round'
        copy_cell_style(ws_tmpl.cell(5, 1), ws['A5'])
        headers = ['Passed', 'Failed', 'Pending', 'N/A']
        for ci, h in enumerate(headers, start=2):
            cell = ws.cell(5, ci)
            cell.value = h
            copy_cell_style(ws_tmpl.cell(5, ci), cell)

        # Row 6: Round 1
        ws['A6'] = 'Round 1'
        copy_cell_style(ws_tmpl.cell(6, 1), ws['A6'])
        ws['B6'] = '=COUNTIF($F$12:$F$5000, "Passed")'
        ws['C6'] = '=COUNTIF($F$12:$F$5000, "Failed")'
        ws['D6'] = '=COUNTIF($F$12:$F$5000, "Pending")'
        ws['E6'] = '=COUNTIF($F$12:$F$5000, "N/A")'
        for ci in range(2, 6):
            copy_cell_style(ws_tmpl.cell(6, ci), ws.cell(6, ci))

        # Row 7: Round 2
        ws['A7'] = 'Round 2'
        copy_cell_style(ws_tmpl.cell(7, 1), ws['A7'])
        ws['B7'] = '=COUNTIF($I$12:$I$5000, "Passed")'
        ws['C7'] = '=COUNTIF($I$12:$I$5000, "Failed")'
        ws['D7'] = '=COUNTIF($I$12:$I$5000, "Pending")'
        ws['E7'] = '=COUNTIF($I$12:$I$5000, "N/A")'
        for ci in range(2, 6):
            copy_cell_style(ws_tmpl.cell(7, ci), ws.cell(7, ci))

        # Row 8: Round 3
        ws['A8'] = 'Round 3'
        copy_cell_style(ws_tmpl.cell(8, 1), ws['A8'])
        ws['B8'] = '=COUNTIF($L$12:$L$5000, "Passed")'
        ws['C8'] = '=COUNTIF($L$12:$L$5000, "Failed")'
        ws['D8'] = '=COUNTIF($L$12:$L$5000, "Pending")'
        ws['E8'] = '=COUNTIF($L$12:$L$5000, "N/A")'
        for ci in range(2, 6):
            copy_cell_style(ws_tmpl.cell(8, ci), ws.cell(8, ci))

        # 2. Row 10 TABLE HEADER: Copy exact text, font, fill, border from template row 10
        table_headers = [
            "Test Case ID", "Test Case Description", "Test Case Procedure",
            "Expected Results", "Pre-conditions",
            "Round 1", "Test date", "Tester",
            "Round 2", "Test date", "Tester",
            "Round 3", "Test date", "Tester",
            "Note"
        ]
        ws.row_dimensions[10].height = ws_tmpl.row_dimensions[10].height or 28.0
        for ci, h in enumerate(table_headers, start=1):
            cell = ws.cell(10, ci)
            cell.value = h
            copy_cell_style(ws_tmpl.cell(10, ci), cell)

        # 3. Clear old rows 11 downwards
        for r in range(11, max(ws.max_row, 30) + 1):
            for c in range(1, 16):
                ws.cell(row=r, column=c).value = None

        # 4. Fill Scenario and Data Rows starting at Row 11
        tcs = get_tcs_for_module(mod, surefire)
        current_row = 11
        current_scenario = None

        for tc in tcs:
            if tc['scenario'] != current_scenario:
                current_scenario = tc['scenario']
                ws.row_dimensions[current_row].height = 20.0
                sc_label = f"Scenario: {current_scenario}"
                ws.cell(current_row, 1).value = sc_label
                # Copy Scenario style from template Row 11 across all 15 cols
                for ci in range(1, 16):
                    cell = ws.cell(current_row, ci)
                    copy_cell_style(ws_tmpl.cell(11, ci), cell)

                current_row += 1

            r = current_row
            ws.cell(row=r, column=1).value = tc['id']
            ws.cell(row=r, column=2).value = tc['desc']
            ws.cell(row=r, column=3).value = tc['proc']
            ws.cell(row=r, column=4).value = tc['exp']
            ws.cell(row=r, column=5).value = tc['pre']

            r1_st = tc['r1']
            c6 = ws.cell(row=r, column=6)
            c6.value = r1_st
            if r1_st in status_fills:
                c6.fill = status_fills[r1_st]
            c6.alignment = data_align_center

            ws.cell(row=r, column=7).value = tc['dt']
            ws.cell(row=r, column=7).alignment = data_align_center
            ws.cell(row=r, column=8).value = tc['tester']
            ws.cell(row=r, column=8).alignment = data_align_center

            c9 = ws.cell(row=r, column=9)
            c9.value = 'Pending'
            c9.fill = status_fills['Pending']
            c9.alignment = data_align_center

            ws.cell(row=r, column=10).value = ''
            ws.cell(row=r, column=11).value = ''

            c12 = ws.cell(row=r, column=12)
            c12.value = 'Pending'
            c12.fill = status_fills['Pending']
            c12.alignment = data_align_center

            ws.cell(row=r, column=13).value = ''
            ws.cell(row=r, column=14).value = ''
            ws.cell(row=r, column=15).value = tc['note']

            # Apply data row styles & borders
            for c in range(1, 16):
                cell = ws.cell(row=r, column=c)
                if c not in (6, 9, 12):
                    copy_cell_style(ws_tmpl.cell(12, c), cell)
                else:
                    cell.font = tahoma_regular
                    cell.border = thin_border
                if c in (1, 2, 3, 4, 5, 15):
                    cell.alignment = data_align_wrap

            max_lines = max(
                len(str(tc['proc']).split('\n')),
                len(str(tc['exp']).split('\n')),
                len(str(tc['desc'])) // 30 + 1
            )
            ws.row_dimensions[r].height = max(25, max_lines * 16)
            current_row += 1

    wb_out.save(str(EXCEL_OUT))
    print(f"Successfully updated all Workflow sheets in {EXCEL_OUT}")

if __name__ == '__main__':
    fix_workflow_sheets()
