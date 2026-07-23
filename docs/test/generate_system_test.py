# -*- coding: utf-8 -*-
"""
WMS System Test Excel Generator
================================
Strategy (confirmed via clarification):
  Q1: Procedure  → Auto-extract method name → AI-style business steps
  Q2: Pre-cond   → @BeforeEach + .sdd/ spec context
  Q3: Round 1    → Maven Surefire XML + Jest JSON (Round 2-3 = manual UAT)
  Q4: Granularity → Group by business scenario (~150-200 TCs total)
  Q5: Output     → .xlsx auto-generated

Usage:
    python docs/test/generate_system_test.py
Output:
    docs/test/SystemTest_WMS_<date>.xlsx
"""

import os
import re
import sys
import xml.etree.ElementTree as ET
from datetime import datetime, date
from pathlib import Path
from collections import defaultdict

# --- openpyxl dependency check ---
import sys
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURATION
# ============================================================
ROOT = Path(r"d:\swp\Manager-warehouse-sdd")
SUREFIRE_DIR = ROOT / "backend" / "target" / "surefire-reports"
BACKEND_TEST_DIR = ROOT / "backend" / "src" / "test" / "java" / "com" / "wms"
FRONTEND_TEST_DIRS = [
    ROOT / "frontend" / "src",
    ROOT / "frontend" / "test_frontend",
]
SPEC_DIR = ROOT / ".sdd" / "specs"
OUT_DIR = ROOT / "docs" / "test"
RUN_DATE = datetime.now().strftime("%Y-%m-%d")
TESTER_AUTO = "CI/Automated"

# ============================================================
# MODULE MAPPING: spec folder → sheet name + code prefix + class keywords
# ============================================================
MODULES = [
    {
        "id": "001",
        "name": "Security, Auth & RBAC",
        "short": "AUTH",
        "spec_dir": "001-security-auth-rbac-audit",
        "be_keywords": ["Auth", "Security", "User", "AuditLog", "SystemConfig", "WarehouseIsolation", "RBAC"],
        "fe_keywords": ["rbac", "config", "admin"],
        "scenarios": [
            ("AUTH-01", "Đăng nhập hệ thống", "Xác thực tài khoản người dùng với các role khác nhau"),
            ("AUTH-02", "Làm mới token (Refresh Token)", "Refresh token hợp lệ / hết hạn"),
            ("AUTH-03", "Đổi mật khẩu", "Đổi mật khẩu đúng / sai mật khẩu cũ / yếu"),
            ("AUTH-04", "Quản lý người dùng", "Tạo, cập nhật, vô hiệu hóa tài khoản"),
            ("AUTH-05", "Phân quyền theo Role", "ADMIN/CEO toàn quyền; WAREHOUSE_MANAGER giới hạn kho"),
            ("AUTH-06", "Kiểm soát truy cập kho (Warehouse Isolation)", "User chỉ thấy dữ liệu kho được gán"),
            ("AUTH-07", "Cấu hình hệ thống (System Config)", "Cập nhật các tham số hệ thống với boundary validation"),
            ("AUTH-08", "Audit Log", "Mọi thao tác nghiệp vụ tạo audit record đầy đủ"),
        ],
    },
    {
        "id": "002",
        "name": "Master Data Management",
        "short": "MDM",
        "spec_dir": "002-master-data-management",
        "be_keywords": ["Product", "Warehouse", "WarehouseLocation", "Supplier", "Dealer", "Vehicle", "Driver"],
        "fe_keywords": ["ProductManagement", "masterData"],
        "scenarios": [
            ("MDM-01", "Quản lý sản phẩm (Product)", "Tạo / cập nhật / tìm kiếm / vô hiệu hóa sản phẩm"),
            ("MDM-02", "Quản lý kho (Warehouse)", "Tạo kho vật lý / In-Transit; vô hiệu hóa khi còn tồn kho"),
            ("MDM-03", "Quản lý vị trí kho (Zone & Bin)", "Tạo zone / bin; kiểm tra capacity khi thu hẹp"),
            ("MDM-04", "Quản lý nhà cung cấp (Supplier)", "Tạo / cập nhật / tìm kiếm nhà cung cấp"),
            ("MDM-05", "Quản lý đại lý (Dealer)", "Tạo đại lý; quản lý credit status lock/unlock"),
            ("MDM-06", "Quản lý xe (Vehicle)", "Tạo xe; vô hiệu hóa xe đang trong chuyến"),
            ("MDM-07", "Quản lý tài xế (Driver)", "Tạo tài xế; gán kho; vô hiệu hóa đang chạy"),
        ],
    },
    {
        "id": "003",
        "name": "Inbound Receipt & QC",
        "short": "RCV",
        "spec_dir": "003-inbound-receipt-qc",
        "be_keywords": ["Receipt", "ReceiptQc", "ReceiptPutaway", "Quarantine", "InboundReceipt", "ReceiptValidation", "ReceiptBackend", "ReceiptRtv"],
        "fe_keywords": ["ReceiptList", "QuarantineWorkspace", "inbound"],
        "scenarios": [
            ("RCV-01", "Tạo phiếu nhập kho", "Tạo phiếu nhập với thông tin NCC, sản phẩm, số lượng"),
            ("RCV-02", "Phê duyệt phiếu nhập", "Approval flow: DRAFT → APPROVED → RECEIVING"),
            ("RCV-03", "Kiểm tra chất lượng (QC Inbound)", "QC Pass / QC Fail từng dòng hàng"),
            ("RCV-04", "Putaway hàng vào Bin", "Gán hàng đã QC pass vào bin; kiểm tra capacity"),
            ("RCV-05", "Xử lý hàng Quarantine", "Hàng fail QC → quarantine zone; không tính available"),
            ("RCV-06", "Trả hàng NCC (RTV)", "Tạo phiếu RTV từ quarantine; xác nhận trả hàng"),
            ("RCV-07", "Tìm kiếm & lọc phiếu nhập", "Filter theo trạng thái, kho, ngày tháng"),
        ],
    },
    {
        "id": "004",
        "name": "Outbound Delivery & POD",
        "short": "OUT",
        "spec_dir": "004-outbound-delivery-pod",
        "be_keywords": ["DeliveryOrder", "Trip", "DriverDelivery", "TripService", "TripController"],
        "fe_keywords": ["DeliveryOrders", "QCOutbound", "DriverTrip", "outbound"],
        "scenarios": [
            ("OUT-01", "Tạo Delivery Order (DO)", "Tạo DO cho đại lý với danh sách sản phẩm và số lượng"),
            ("OUT-02", "QC Outbound (kiểm hàng xuất)", "Kiểm tra hàng trước khi xuất kho theo FIFO"),
            ("OUT-03", "Tạo chuyến giao hàng (Trip)", "Gán DO vào Trip; kiểm tra trọng tải xe"),
            ("OUT-04", "Xuất phát chuyến (Depart Trip)", "Depart: kiểm tra QC pass record đủ số lượng"),
            ("OUT-05", "Xác nhận giao hàng (POD)", "Driver xác nhận giao; cập nhật trạng thái DO"),
            ("OUT-06", "Hủy / Từ chối chuyến", "Hủy Trip ở trạng thái PLANNED; release xe và tài xế"),
            ("OUT-07", "Hoàn thành chuyến (Complete Trip)", "Complete khi tất cả DO terminal; release vehicle/driver"),
        ],
    },
    {
        "id": "005",
        "name": "Inter-Warehouse Transfer",
        "short": "TRF",
        "spec_dir": "005-inter-warehouse-transfer",
        "be_keywords": ["TransferRequest", "Transfer", "InterWarehouseTransfer"],
        "fe_keywords": ["TransferRequest", "inter-warehouse-transfer", "validation"],
        "scenarios": [
            ("TRF-01", "Tạo yêu cầu điều chuyển", "Tạo transfer request với source/dest khác nhau"),
            ("TRF-02", "Submit & Approve yêu cầu", "Submit bởi WH Manager; CEO Approve/Reject"),
            ("TRF-03", "Chuyển thành Transfer thực tế", "Planner convert request thành transfer"),
            ("TRF-04", "QC Outbound tại kho nguồn", "Kiểm hàng trước khi xuất; worker load report"),
            ("TRF-05", "Bàn giao hàng cho tài xế", "Load handover photo; chuyển sang In-Transit"),
            ("TRF-06", "Nhận hàng tại kho đích", "Receiving handover; đếm số lượng thực nhận"),
            ("TRF-07", "QC nhận hàng tại kho đích", "QC pass → putaway; QC fail → quarantine"),
            ("TRF-08", "Xử lý hàng trả về nguồn (Return)", "Wrong-SKU / partial → return to source flow"),
            ("TRF-09", "Chênh lệch số lượng gửi/nhận", "Variance → tạo adjustment/audit record"),
        ],
    },
    {
        "id": "006",
        "name": "Stocktake & Adjustment",
        "short": "STK",
        "spec_dir": "006-stocktake-adjustment",
        "be_keywords": ["StockTake", "Adjustment"],
        "fe_keywords": ["StocktakeList", "stocktake"],
        "scenarios": [
            ("STK-01", "Tạo phiếu kiểm kê", "Tạo stocktake cho kho / zone cụ thể"),
            ("STK-02", "Nhập số lượng thực đếm", "Nhân viên nhập số lượng từng bin"),
            ("STK-03", "Submit phiếu kiểm kê", "Submit để tính variance"),
            ("STK-04", "Phê duyệt chênh lệch (Variance Approval)", "CEO / Manager duyệt variance"),
            ("STK-05", "Điều chỉnh tồn kho (Adjustment)", "Tạo adjustment từ variance đã duyệt"),
            ("STK-06", "Lọc & tìm kiếm phiếu kiểm kê", "Filter theo trạng thái, kho, ngày"),
        ],
    },
    {
        "id": "007",
        "name": "Pricing & COGS",
        "short": "PRC",
        "spec_dir": "007-pricing-cogs-management",
        "be_keywords": ["PriceHistory", "Price"],
        "fe_keywords": ["pricing"],
        "scenarios": [
            ("PRC-01", "Tạo bảng giá (Price History)", "Tạo giá bán cho sản phẩm theo ngày hiệu lực"),
            ("PRC-02", "Phê duyệt giá", "CEO approve price history entry"),
            ("PRC-03", "Xem lịch sử giá", "Filter price history theo sản phẩm, ngày"),
            ("PRC-04", "Tính COGS (Cost of Goods Sold)", "COGS tự động theo FIFO khi xuất hàng"),
        ],
    },
    {
        "id": "008",
        "name": "Finance, Billing & Closing",
        "short": "FIN",
        "spec_dir": "008-finance-billing-closing",
        "be_keywords": ["Invoice", "Payment", "AccountingPeriod", "AutoInvoice", "SupplierInvoice", "SupplierPayment", "BillingNotification"],
        "fe_keywords": ["PeriodClosing", "SupplierInvoices", "DealerDebtInvoice", "PaymentReceipts", "finance"],
        "scenarios": [
            ("FIN-01", "Tạo hóa đơn tự động (Auto Invoice)", "Invoice tự động khi delivery xác nhận hoàn thành"),
            ("FIN-02", "Hóa đơn NCC (Supplier Invoice)", "Tạo / xem hóa đơn từ nhà cung cấp"),
            ("FIN-03", "Thanh toán NCC (Supplier Payment)", "Ghi nhận thanh toán; cập nhật công nợ"),
            ("FIN-04", "Quản lý công nợ đại lý", "Xem nợ; credit hold khi vượt hạn mức"),
            ("FIN-05", "Đóng kỳ kế toán (Period Closing)", "Mở / đóng kỳ; không cho phép giao dịch sau đóng"),
            ("FIN-06", "Báo cáo công nợ (Credit Aging)", "Xem aging report theo đại lý"),
        ],
    },
    {
        "id": "009",
        "name": "Returns, Scrap & Disposal",
        "short": "RET",
        "spec_dir": "009-returns-scrap-disposal",
        "be_keywords": ["Returns", "Disposal", "ReceiptRtv", "ReceiptServiceReturn"],
        "fe_keywords": ["returns"],
        "scenarios": [
            ("RET-01", "Tạo phiếu trả hàng khách (Customer Return)", "Khách trả hàng; tạo receipt hoàn trả"),
            ("RET-02", "QC hàng trả", "Kiểm tra chất lượng hàng trả; pass/fail"),
            ("RET-03", "Cấp Credit Note", "Phê duyệt và cấp credit note cho đại lý"),
            ("RET-04", "Xử lý hàng tiêu hủy (Disposal)", "Tạo phiếu tiêu hủy; duyệt; cập nhật tồn kho"),
        ],
    },
    {
        "id": "010",
        "name": "Reports, Dashboard & Alerts",
        "short": "RPT",
        "spec_dir": "010-reports-dashboards-alerts",
        "be_keywords": ["Report", "StockAlert", "ReportService"],
        "fe_keywords": ["report", "CreditAgingReport"],
        "scenarios": [
            ("RPT-01", "CEO Dashboard", "Xem KPI tổng hợp: doanh thu, tồn kho, chuyến giao"),
            ("RPT-02", "Báo cáo định giá tồn kho", "Inventory Valuation theo FIFO"),
            ("RPT-03", "Cảnh báo tồn kho thấp (Low Stock Alert)", "Alert khi dưới reorder point; resolve khi đủ hàng"),
            ("RPT-04", "Báo cáo công nợ theo độ tuổi", "Credit aging report phân nhóm 30/60/90 ngày"),
        ],
    },
]

# ============================================================
# STEP 1: Parse Surefire XML → {classname: {method: status}}
# ============================================================
def parse_surefire_reports(surefire_dir: Path) -> dict:
    results = {}  # classname → {method → {status, time, timestamp}}
    if not surefire_dir.exists():
        return results

    suite_time = datetime.now().strftime("%Y-%m-%d")
    for xml_file in surefire_dir.glob("TEST-*.xml"):
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()
            suite_name = root.get("name", "")
            # Try to get timestamp from suite
            ts = root.get("timestamp", suite_time)
            if ts and "T" in ts:
                ts = ts.split("T")[0]

            for tc in root.findall("testcase"):
                method = tc.get("name", "")
                classname = tc.get("classname", suite_name)
                elapsed = tc.get("time", "0")

                if tc.find("failure") is not None or tc.find("error") is not None:
                    status = "Failed"
                elif tc.find("skipped") is not None:
                    status = "N/A"
                else:
                    status = "Passed"

                short_class = classname.split(".")[-1]
                if short_class not in results:
                    results[short_class] = {}
                results[short_class][method] = {"status": status, "date": ts, "time": elapsed}
        except Exception:
            continue
    return results


# ============================================================
# STEP 2: Parse Java test files → extract methods + BeforeEach
# ============================================================
def snake_to_words(name: str) -> str:
    """Convert camelCase/snake_case method name to readable words."""
    # Split on underscore first
    parts = name.split("_")
    words = []
    for part in parts:
        # Split camelCase
        sub = re.sub(r"([A-Z])", r" \1", part).strip()
        words.append(sub.lower())
    return " ".join(words)


def method_to_procedure(method_name: str) -> str:
    """Convert test method name to business procedure steps."""
    words = snake_to_words(method_name)

    # Pattern-based step generation
    steps = []
    w = words.lower()

    # Identify actor/context
    if any(x in w for x in ["login", "auth", "signin"]):
        steps.append("1. Mở trình duyệt, truy cập trang đăng nhập")
        steps.append("2. Nhập thông tin tài khoản")
        steps.append("3. Nhấn nút Đăng nhập")
    elif any(x in w for x in ["create", "tạo", "add"]):
        steps.append("1. Đăng nhập với role phù hợp")
        steps.append("2. Điều hướng đến chức năng tương ứng")
        steps.append("3. Nhấn nút Tạo mới / New")
        steps.append("4. Điền đầy đủ thông tin theo yêu cầu")
        steps.append("5. Nhấn Lưu / Submit")
    elif any(x in w for x in ["update", "edit", "sửa", "cập nhật"]):
        steps.append("1. Đăng nhập với role phù hợp")
        steps.append("2. Tìm và mở bản ghi cần cập nhật")
        steps.append("3. Nhấn Chỉnh sửa / Edit")
        steps.append("4. Thay đổi thông tin cần thiết")
        steps.append("5. Nhấn Lưu")
    elif any(x in w for x in ["delete", "deactivate", "disable", "vô hiệu"]):
        steps.append("1. Đăng nhập với role phù hợp")
        steps.append("2. Tìm bản ghi cần xử lý")
        steps.append("3. Nhấn Vô hiệu hóa / Xóa")
        steps.append("4. Xác nhận thao tác trong hộp thoại")
    elif any(x in w for x in ["approve", "phê duyệt", "duyệt"]):
        steps.append("1. Đăng nhập với role có quyền phê duyệt")
        steps.append("2. Mở danh sách chờ duyệt")
        steps.append("3. Chọn phiếu cần duyệt")
        steps.append("4. Xem xét thông tin và nhấn Phê duyệt")
    elif any(x in w for x in ["reject", "từ chối"]):
        steps.append("1. Đăng nhập với role có quyền từ chối")
        steps.append("2. Mở danh sách chờ xử lý")
        steps.append("3. Chọn phiếu và nhấn Từ chối")
        steps.append("4. Nhập lý do từ chối")
    elif any(x in w for x in ["list", "get all", "search", "filter", "tìm"]):
        steps.append("1. Đăng nhập hệ thống")
        steps.append("2. Điều hướng đến danh sách")
        steps.append("3. Áp dụng bộ lọc nếu cần")
        steps.append("4. Quan sát kết quả trả về")
    elif any(x in w for x in ["cancel", "hủy"]):
        steps.append("1. Đăng nhập với role phù hợp")
        steps.append("2. Mở phiếu cần hủy")
        steps.append("3. Nhấn Hủy / Cancel")
        steps.append("4. Xác nhận hủy trong hộp thoại")
    else:
        steps.append("1. Đăng nhập hệ thống với role phù hợp")
        steps.append("2. Thực hiện thao tác: " + words[:50])
        steps.append("3. Quan sát kết quả")

    # Append error case note if relevant
    if any(x in w for x in ["rejects", "fails", "invalid", "error", "throws", "exception", "returns 4"]):
        steps.append(f"[Negative] Kiểm tra: {words[:60]}")

    return "\n".join(steps)


def method_to_expected(method_name: str) -> str:
    """Derive expected result from test method name."""
    w = method_name.lower()

    if any(x in w for x in ["success", "valid", "returns200", "returns201", "returns204"]):
        return "Hệ thống xử lý thành công.\nThông báo success hiển thị.\nDữ liệu được lưu đúng."
    elif any(x in w for x in ["rejects", "fails", "returns400", "throws", "invalid", "error"]):
        return "Hệ thống hiển thị thông báo lỗi phù hợp.\nKhông có dữ liệu nào bị thay đổi.\nHTTP status 4xx được trả về."
    elif any(x in w for x in ["returns401", "unauthenticated", "unauthorized"]):
        return "Hệ thống từ chối truy cập.\nHTTP 401 Unauthorized.\nYêu cầu đăng nhập lại."
    elif any(x in w for x in ["returns403", "forbidden", "notallowed"]):
        return "Hệ thống từ chối quyền truy cập.\nHTTP 403 Forbidden.\nThông báo không có quyền hiển thị."
    elif any(x in w for x in ["notfound", "returns404"]):
        return "Hệ thống trả về thông báo không tìm thấy.\nHTTP 404 Not Found."
    elif any(x in w for x in ["list", "getall", "search"]):
        return "Danh sách dữ liệu hiển thị đúng.\nCác bộ lọc hoạt động chính xác."
    else:
        return "Hệ thống phản hồi đúng theo nghiệp vụ.\nDữ liệu hiển thị chính xác."


def extract_before_each(java_content: str) -> str:
    """Extract @BeforeEach setup to derive pre-conditions."""
    pattern = r"@BeforeEach[\s\S]*?void\s+\w+\s*\(\s*\)\s*\{([\s\S]*?)(?=\n\s*@|\n\s*@Test|\n\s*}[\s]*\n\s*@)"
    match = re.search(pattern, java_content)
    if not match:
        return ""

    body = match.group(1)
    preconditions = []

    # Extract entity setup patterns
    if re.search(r"new\s+User|mockUser|adminUser|managerUser", body):
        preconditions.append("Tài khoản người dùng đã được tạo trong hệ thống")
    if re.search(r"new\s+Warehouse|mockWarehouse|warehouse\s*=", body):
        preconditions.append("Kho hàng đã được khởi tạo và kích hoạt")
    if re.search(r"new\s+Product|mockProduct|product\s*=", body):
        preconditions.append("Sản phẩm đã được đăng ký trong hệ thống")
    if re.search(r"new\s+Supplier|mockSupplier|supplier\s*=", body):
        preconditions.append("Nhà cung cấp đã được đăng ký")
    if re.search(r"new\s+Receipt|receipt\s*=|mockReceipt", body):
        preconditions.append("Phiếu nhập kho đã tồn tại trong hệ thống")
    if re.search(r"new\s+Inventory|inventory\s*=|mockInventory", body):
        preconditions.append("Tồn kho hiện tại đã được thiết lập")
    if re.search(r"JWT|token|jwtToken|authHeader", body):
        preconditions.append("Token xác thực hợp lệ đã được cấp phát")
    if re.search(r"new\s+Trip|trip\s*=|mockTrip", body):
        preconditions.append("Chuyến giao hàng đã được khởi tạo")
    if re.search(r"new\s+DeliveryOrder|deliveryOrder\s*=", body):
        preconditions.append("Delivery Order đã được tạo và phê duyệt")

    if not preconditions:
        preconditions.append("Hệ thống đang hoạt động bình thường")
        preconditions.append("Người dùng đã đăng nhập với role phù hợp")

    return "\n".join(f"- {p}" for p in preconditions)


def parse_java_tests(be_test_dir: Path, surefire_results: dict, module_keywords: list) -> list:
    """Parse Java test files for a module. Returns list of TC dicts."""
    tcs = []
    for root, _, files in os.walk(be_test_dir):
        for f in files:
            if not f.endswith(".java"):
                continue
            class_name = f.replace(".java", "")

            # Match to module keywords
            matched = any(kw.lower() in class_name.lower() for kw in module_keywords)
            if not matched:
                continue

            path = Path(root) / f
            try:
                content = path.read_text(encoding="utf-8", errors="ignore")
            except Exception:
                continue

            # Extract @BeforeEach for pre-conditions
            preconditions = extract_before_each(content)
            if not preconditions:
                preconditions = "- Hệ thống đang hoạt động bình thường\n- Người dùng đã đăng nhập với role phù hợp"

            # Extract test methods
            methods = re.findall(r"@Test[\s\S]{0,300}?void\s+(\w+)\s*\(", content)
            class_results = surefire_results.get(class_name, {})

            for method in methods:
                result = class_results.get(method, {})
                status = result.get("status", "Pending")
                run_date = result.get("date", "")

                tcs.append({
                    "id": f"TC-{method[:6].upper()}",
                    "description": snake_to_words(method).title(),
                    "class": class_name,
                    "method": method,
                    "procedure": method_to_procedure(method),
                    "expected": method_to_expected(method),
                    "preconditions": preconditions,
                    "round1_status": status,
                    "round1_date": run_date,
                    "round1_tester": TESTER_AUTO if status in ("Passed", "Failed") else "",
                    "note": f"Source: {class_name}.java",
                })
    return tcs


def parse_frontend_tests(fe_dirs: list, module_keywords: list) -> list:
    """Parse frontend test files for a module."""
    tcs = []
    pattern = re.compile(r"""(?:it|test)\s*\(['"](.*?)['"]\s*,""")

    for base in fe_dirs:
        if not Path(base).exists():
            continue
        for root, _, files in os.walk(base):
            for f in files:
                if ".test." not in f:
                    continue
                matched = any(kw.lower() in f.lower() for kw in module_keywords)
                if not matched:
                    continue

                path = Path(root) / f
                try:
                    content = path.read_text(encoding="utf-8", errors="ignore")
                except Exception:
                    continue

                matches = pattern.findall(content)
                for m in matches:
                    tcs.append({
                        "id": f"TC-FE-{m[:8].upper().replace(' ', '_')}",
                        "description": m,
                        "class": f,
                        "method": m,
                        "procedure": method_to_procedure(m),
                        "expected": method_to_expected(m),
                        "preconditions": "- Trình duyệt đã mở ứng dụng WMS\n- Người dùng đã đăng nhập",
                        "round1_status": "Pending",
                        "round1_date": "",
                        "round1_tester": "",
                        "note": f"Frontend: {f}",
                    })
    return tcs


# ============================================================
# STEP 3: Group TCs into scenarios
# ============================================================
def group_into_scenarios(raw_tcs: list, scenarios: list) -> list:
    """
    Group raw TCs into business scenarios.
    Returns list of scenario groups: {scenario_id, name, desc, tcs: []}
    """
    # Build keyword sets per scenario
    scenario_groups = []
    for s_id, s_name, s_desc in scenarios:
        kws = s_name.lower().split() + s_desc.lower().split()
        scenario_groups.append({
            "id": s_id,
            "name": s_name,
            "desc": s_desc,
            "keywords": set(kws),
            "tcs": [],
        })

    unmatched = []
    for tc in raw_tcs:
        words = set((tc["description"] + " " + tc["method"]).lower().split("_"))
        best_score = 0
        best_idx = -1
        for i, sg in enumerate(scenario_groups):
            score = len(words & sg["keywords"])
            if score > best_score:
                best_score = score
                best_idx = i
        if best_idx >= 0 and best_score > 0:
            scenario_groups[best_idx]["tcs"].append(tc)
        else:
            unmatched.append(tc)

    # Put unmatched into first scenario
    if unmatched and scenario_groups:
        scenario_groups[0]["tcs"].extend(unmatched)

    return scenario_groups


# ============================================================
# STEP 4: Excel styles
# ============================================================
COLORS = {
    "header_bg": "1F3864",       # Navy blue
    "header_font": "FFFFFF",
    "scenario_bg": "2E75B6",     # Mid blue
    "scenario_font": "FFFFFF",
    "alt_row": "DCE6F1",         # Light blue
    "passed_bg": "C6EFCE",       # Green
    "failed_bg": "FFC7CE",       # Red
    "pending_bg": "FFEB9C",      # Yellow
    "na_bg": "D9D9D9",           # Gray
    "cover_title": "1F3864",
    "cover_accent": "2E75B6",
}

COL_WIDTHS = {
    "A": 18,   # TC ID
    "B": 40,   # Description
    "C": 50,   # Procedure
    "D": 40,   # Expected
    "E": 35,   # Pre-conditions
    "F": 12,   # Round 1 Status
    "G": 14,   # Round 1 Date
    "H": 18,   # Round 1 Tester
    "I": 12,   # Round 2 Status
    "J": 14,   # Round 2 Date
    "K": 18,   # Round 2 Tester
    "L": 12,   # Round 3 Status
    "M": 14,   # Round 3 Date
    "N": 18,   # Round 3 Tester
    "O": 30,   # Note
}

STATUS_COLORS = {
    "Passed": "C6EFCE",
    "Failed": "FFC7CE",
    "Pending": "FFEB9C",
    "N/A": "D9D9D9",
}

def make_border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def header_font():
    return Font(name="Calibri", bold=True, color=COLORS["header_font"], size=10)

def cell_font(bold=False):
    return Font(name="Calibri", bold=bold, size=9)

def wrap_align(horizontal="left", vertical="top"):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=True)


def style_header_cell(cell, text, bg_color=None):
    bg = bg_color or COLORS["header_bg"]
    cell.value = text
    cell.font = header_font()
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = wrap_align("center", "center")
    cell.border = make_border()


def style_scenario_row(cell, text):
    cell.value = text
    cell.font = Font(name="Calibri", bold=True, color=COLORS["scenario_font"], size=9)
    cell.fill = PatternFill("solid", fgColor=COLORS["scenario_bg"])
    cell.alignment = wrap_align("left", "center")
    cell.border = make_border()


def style_data_cell(cell, text, row_alt=False, status=None):
    cell.value = text
    cell.font = cell_font()
    cell.alignment = wrap_align()
    cell.border = make_border()

    if status and status in STATUS_COLORS:
        cell.fill = PatternFill("solid", fgColor=STATUS_COLORS[status])
    elif row_alt:
        cell.fill = PatternFill("solid", fgColor=COLORS["alt_row"])


# ============================================================
# STEP 5: Build Cover Sheet
# ============================================================
def build_cover_sheet(wb, modules_summary):
    ws = wb.active
    ws.title = "Cover"

    # Title
    ws.merge_cells("B2:H2")
    title_cell = ws["B2"]
    title_cell.value = "SYSTEM TEST REPORT — WMS (Warehouse Management System)"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color=COLORS["cover_title"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Meta info
    meta = [
        ("Project", "WMS — Warehouse Management System"),
        ("Version", "Sprint 1 — Release 1.0"),
        ("Test Date", RUN_DATE),
        ("Modules", "10 modules (Specs 001–010)"),
        ("Notes", "Round 1: Automated (Maven Surefire + Jest). Round 2-3: Manual UAT on staging."),
    ]
    for i, (k, v) in enumerate(meta, start=4):
        row = i
        kc = ws.cell(row=row, column=2, value=k)
        kc.font = Font(name="Calibri", bold=True, size=10, color="1F3864")
        kc.alignment = Alignment(horizontal="right", vertical="center")
        vc = ws.cell(row=row, column=3, value=v)
        vc.font = Font(name="Calibri", size=10)
        vc.alignment = Alignment(horizontal="left", vertical="center")

    # Summary table header
    header_row = 11
    headers = ["No", "Module", "Module Code", "Passed", "Failed", "Pending", "N/A", "Total TCs"]
    for ci, h in enumerate(headers, start=2):
        c = ws.cell(row=header_row, column=ci, value=h)
        c.font = header_font()
        c.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
        c.alignment = wrap_align("center", "center")
        c.border = make_border()

    # Summary rows
    for ri, mod in enumerate(modules_summary, start=1):
        row = header_row + ri
        vals = [
            ri,
            mod["name"],
            mod["short"],
            mod["passed"],
            mod["failed"],
            mod["pending"],
            mod["na"],
            mod["total"],
        ]
        for ci, v in enumerate(vals, start=2):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = cell_font()
            c.border = make_border()
            c.alignment = Alignment(horizontal="center" if ci > 4 else "left", vertical="center")

    # Totals row
    total_row = header_row + len(modules_summary) + 1
    total_vals = ["", "TOTAL", "",
                  sum(m["passed"] for m in modules_summary),
                  sum(m["failed"] for m in modules_summary),
                  sum(m["pending"] for m in modules_summary),
                  sum(m["na"] for m in modules_summary),
                  sum(m["total"] for m in modules_summary)]
    for ci, v in enumerate(total_vals, start=2):
        c = ws.cell(row=total_row, column=ci, value=v)
        c.font = Font(name="Calibri", bold=True, size=10)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
        c.border = make_border()
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Coverage formulas (approximate)
    cov_row = total_row + 2
    ws.cell(row=cov_row, column=2, value="Test Coverage").font = Font(name="Calibri", bold=True, size=10)
    total_tc = sum(m["total"] for m in modules_summary)
    total_na = sum(m["na"] for m in modules_summary)
    total_done = sum(m["passed"] + m["failed"] for m in modules_summary)
    denom = total_tc - total_na or 1
    cov_pct = round(total_done * 100 / denom, 1)
    ws.cell(row=cov_row, column=3, value=f"{cov_pct}%").font = Font(name="Calibri", size=10, bold=True, color="2E75B6")

    pass_row = cov_row + 1
    ws.cell(row=pass_row, column=2, value="Pass Rate").font = Font(name="Calibri", bold=True, size=10)
    total_pass = sum(m["passed"] for m in modules_summary)
    pass_pct = round(total_pass * 100 / denom, 1)
    ws.cell(row=pass_row, column=3, value=f"{pass_pct}%").font = Font(name="Calibri", size=10, bold=True, color="375623")

    # Column widths
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 16
    for col in ["E", "F", "G", "H", "I"]:
        ws.column_dimensions[col].width = 12

    ws.row_dimensions[2].height = 36


# ============================================================
# STEP 6: Build Module Worksheet
# ============================================================
def build_module_sheet(wb, module: dict, scenario_groups: list) -> dict:
    """Build one worksheet for a module. Returns summary stats."""
    short = module["short"]
    name = module["name"][:28]
    ws = wb.create_sheet(title=f"{short} - {name}"[:31])

    # --- Module header block ---
    ws.merge_cells("A1:O1")
    title = ws["A1"]
    title.value = f"MODULE: {module['id']} — {module['name']}"
    title.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor=COLORS["cover_title"])
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Module meta (rows 2-5)
    meta_pairs = [
        ("Workflow", module["name"]),
        ("Test requirement", f"Kiểm tra toàn bộ nghiệp vụ thuộc module {module['name']} theo spec {module['id']}"),
        ("Number of TCs", "=COUNTA(A10:A5000)"),
    ]
    for i, (k, v) in enumerate(meta_pairs, start=2):
        ws.cell(row=i, column=1, value=k).font = Font(name="Calibri", bold=True, size=10, color="1F3864")
        ws.cell(row=i, column=2, value=v).font = Font(name="Calibri", size=10)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=9)

    # Round summary (row 6-8)
    ws.cell(row=6, column=1, value="Testing Round").font = Font(name="Calibri", bold=True, size=9)
    for ci, h in enumerate(["Passed", "Failed", "Pending", "N/A"], start=2):
        ws.cell(row=6, column=ci, value=h).font = Font(name="Calibri", bold=True, size=9, color=COLORS["header_font"])
        ws.cell(row=6, column=ci).fill = PatternFill("solid", fgColor=COLORS["header_bg"])

    for ri, rnd in enumerate(["Round 1", "Round 2", "Round 3"], start=7):
        ws.cell(row=ri, column=1, value=rnd).font = Font(name="Calibri", bold=True, size=9)
        ws.cell(row=ri, column=2, value=f"=COUNTIF($F$10:$F$5000,\"Passed\")").font = cell_font()
        ws.cell(row=ri, column=3, value=f"=COUNTIF($F$10:$F$5000,\"Failed\")").font = cell_font()
        ws.cell(row=ri, column=4, value=f"=COUNTIF($F$10:$F$5000,\"Pending\")").font = cell_font()
        ws.cell(row=ri, column=5, value=f"=COUNTIF($F$10:$F$5000,\"N/A\")").font = cell_font()

    # --- Column headers row 9 ---
    col_headers = [
        "Test Case ID", "Test Case Description", "Test Case Procedure",
        "Expected Results", "Pre-conditions",
        "Round 1", "Test Date", "Tester",
        "Round 2", "Test Date", "Tester",
        "Round 3", "Test Date", "Tester",
        "Note"
    ]
    for ci, h in enumerate(col_headers, start=1):
        style_header_cell(ws.cell(row=9, column=ci), h)

    ws.row_dimensions[9].height = 30
    ws.freeze_panes = "A10"

    # --- Data rows ---
    current_row = 10
    passed_count = failed_count = pending_count = na_count = 0

    for sg in scenario_groups:
        if not sg["tcs"]:
            continue

        # Scenario separator row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=15)
        scenario_cell = ws.cell(row=current_row, column=1)
        style_scenario_row(scenario_cell, f"  [{sg['id']}] {sg['name']}: {sg['desc']}")
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        for ti, tc in enumerate(sg["tcs"]):
            alt = ti % 2 == 1
            r1_status = tc["round1_status"]

            # Count stats
            if r1_status == "Passed":
                passed_count += 1
            elif r1_status == "Failed":
                failed_count += 1
            elif r1_status == "N/A":
                na_count += 1
            else:
                pending_count += 1

            # Generate TC ID with scenario prefix
            tc_id = f"{sg['id']}-{ti+1:03d}"

            row_data = [
                tc_id,
                tc["description"],
                tc["procedure"],
                tc["expected"],
                tc["preconditions"],
                r1_status,
                tc["round1_date"],
                tc["round1_tester"],
                "Pending", "", "",   # Round 2
                "Pending", "", "",   # Round 3
                tc["note"],
            ]

            for ci, val in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=ci, value=val)
                # Status columns get color
                if ci == 6:
                    style_data_cell(cell, val, alt, status=r1_status)
                elif ci in (9, 12):
                    style_data_cell(cell, val, alt, status="Pending")
                else:
                    style_data_cell(cell, val, alt)

            ws.row_dimensions[current_row].height = 60
            current_row += 1

    # Column widths
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    total = passed_count + failed_count + pending_count + na_count
    return {
        "name": module["name"],
        "short": module["short"],
        "passed": passed_count,
        "failed": failed_count,
        "pending": pending_count,
        "na": na_count,
        "total": total,
    }


# ============================================================
# MAIN
# ============================================================
def main():
    print("=" * 60)
    print("WMS System Test Excel Generator")
    print(f"Run date: {RUN_DATE}")
    print("=" * 60)

    # Parse surefire reports
    print("\n[1/4] Parsing Maven Surefire reports...")
    surefire_results = parse_surefire_reports(SUREFIRE_DIR)
    print(f"  -> Found results for {len(surefire_results)} test classes")

    # Create workbook
    wb = openpyxl.Workbook()

    modules_summary = []

    print("\n[2/4] Processing modules...")
    for module in MODULES:
        print(f"  [{module['id']}] {module['name']}")

        # Parse BE tests
        be_tcs = parse_java_tests(BACKEND_TEST_DIR, surefire_results, module["be_keywords"])

        # Parse FE tests
        fe_tcs = parse_frontend_tests(
            [str(d) for d in FRONTEND_TEST_DIRS],
            module["fe_keywords"]
        )

        all_tcs = be_tcs + fe_tcs
        print(f"      {len(be_tcs)} BE tests + {len(fe_tcs)} FE tests = {len(all_tcs)} total")

        # Group into scenarios
        scenario_groups = group_into_scenarios(all_tcs, module["scenarios"])

        # Build sheet
        summary = build_module_sheet(wb, module, scenario_groups)
        modules_summary.append(summary)
        print(f"      Passed={summary['passed']} Failed={summary['failed']} Pending={summary['pending']}")

    # Build cover
    print("\n[3/4] Building cover sheet...")
    build_cover_sheet(wb, modules_summary)

    # Save
    out_path = OUT_DIR / f"SystemTest_WMS_{RUN_DATE.replace('-', '')}.xlsx"
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    print(f"\n[4/4] Saving to: {out_path}")
    wb.save(str(out_path))

    # Summary
    print("\n" + "=" * 60)
    print("DONE!")
    total_tcs = sum(m["total"] for m in modules_summary)
    total_passed = sum(m["passed"] for m in modules_summary)
    total_failed = sum(m["failed"] for m in modules_summary)
    total_pending = sum(m["pending"] for m in modules_summary)
    print(f"Total TCs generated : {total_tcs}")
    print(f"Round 1 - Passed    : {total_passed}")
    print(f"Round 1 - Failed    : {total_failed}")
    print(f"Round 1 - Pending   : {total_pending}")
    denom = total_tcs - sum(m["na"] for m in modules_summary) or 1
    print(f"Coverage            : {round((total_passed + total_failed) * 100 / denom, 1)}%")
    print(f"Pass rate           : {round(total_passed * 100 / denom, 1)}%")
    print(f"\nOutput: {out_path}")
    return str(out_path)


if __name__ == "__main__":
    main()
