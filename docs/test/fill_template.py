# -*- coding: utf-8 -*-
"""
WMS System Test — Fill Template
================================
Approach: Copy Template3_System Test.xlsx as base, then:
  1. Rename sheet "Workflow Name1" -> first module
  2. Clone "Workflow Name2" for remaining modules
  3. Fill data rows (from row 12+) into each module sheet
  4. Update Test Statistics formulas
  5. Update Cover & Test Cases sheets
  
Minimizes changes to template structure/formatting.
"""

import os, re, sys, shutil, copy, xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet

# ============================================================
# CONFIG
# ============================================================
ROOT        = Path(r'd:\swp\Manager-warehouse-sdd')
TEMPLATE    = ROOT / 'docs' / 'test' / 'Template3_System Test.xlsx'
SUREFIRE    = ROOT / 'backend' / 'target' / 'surefire-reports'
BE_TEST_DIR = ROOT / 'backend' / 'src' / 'test' / 'java' / 'com' / 'wms'
FE_TEST_DIRS = [
    ROOT / 'frontend' / 'src',
    ROOT / 'frontend' / 'test_frontend',
]
OUT_DIR     = ROOT / 'docs' / 'test'
RUN_DATE    = datetime.now().strftime('%Y-%m-%d')
TESTER_AUTO = 'CI/Automated'

PROJECT_NAME = 'WMS - Warehouse Management System'
PROJECT_CODE = 'WMS-SPRINT1'

# ============================================================
# MODULE DEFINITIONS
# ============================================================
MODULES = [
    {
        'id': '001', 'name': 'Security, Auth & RBAC', 'short': 'AUTH-001',
        'req': 'Kiem tra xac thuc nguoi dung, phan quyen theo role, kiem soat truy cap kho, cau hinh he thong va audit log',
        'be_keywords': ['Auth', 'Security', 'User', 'AuditLog', 'SystemConfig', 'WarehouseIsolation', 'RBAC'],
        'fe_keywords': ['rbac', 'config', 'admin'],
        'scenarios': [
            'Dang nhap he thong',
            'Lam moi token (Refresh Token)',
            'Doi mat khau',
            'Quan ly nguoi dung (User CRUD)',
            'Phan quyen theo Role (RBAC)',
            'Kiem soat truy cap kho',
            'Cau hinh he thong (System Config)',
            'Audit Log',
        ],
    },
    {
        'id': '002', 'name': 'Master Data Management', 'short': 'MDM-002',
        'req': 'Kiem tra quan ly du lieu danh muc: san pham, kho, vi tri kho, nha cung cap, dai ly, xe, tai xe',
        'be_keywords': ['Product', 'Warehouse', 'WarehouseLocation', 'Supplier', 'Dealer', 'Vehicle', 'Driver'],
        'fe_keywords': ['ProductManagement', 'masterData'],
        'scenarios': ['Quan ly san pham', 'Quan ly kho', 'Quan ly vi tri kho (Zone & Bin)',
                      'Quan ly nha cung cap', 'Quan ly dai ly', 'Quan ly xe', 'Quan ly tai xe'],
    },
    {
        'id': '003', 'name': 'Inbound Receipt & QC', 'short': 'RCV-003',
        'req': 'Kiem tra toan bo quy trinh nhap kho: tao phieu, phe duyet, QC, putaway, quarantine, RTV',
        'be_keywords': ['Receipt', 'ReceiptQc', 'ReceiptPutaway', 'Quarantine', 'InboundReceipt',
                        'ReceiptValidation', 'ReceiptBackend', 'ReceiptRtv'],
        'fe_keywords': ['ReceiptList', 'QuarantineWorkspace', 'inbound'],
        'scenarios': ['Tao phieu nhap kho', 'Phe duyet phieu nhap', 'QC Inbound',
                      'Putaway hang vao Bin', 'Xu ly hang Quarantine', 'Tra hang NCC (RTV)',
                      'Tim kiem & loc phieu nhap'],
    },
    {
        'id': '004', 'name': 'Outbound Delivery & POD', 'short': 'OUT-004',
        'req': 'Kiem tra quy trinh xuat kho: Delivery Order, QC outbound, chuyen giao hang, POD',
        'be_keywords': ['DeliveryOrder', 'Trip', 'DriverDelivery', 'TripService', 'TripController'],
        'fe_keywords': ['DeliveryOrders', 'QCOutbound', 'DriverTrip', 'outbound'],
        'scenarios': ['Tao Delivery Order', 'QC Outbound', 'Tao chuyen giao hang (Trip)',
                      'Xuat phat chuyen (Depart)', 'Xac nhan giao hang (POD)',
                      'Huy chuyen hang', 'Hoan thanh chuyen (Complete)'],
    },
    {
        'id': '005', 'name': 'Inter-Warehouse Transfer', 'short': 'TRF-005',
        'req': 'Kiem tra quy trinh dieu chuyen hang giua cac kho: yeu cau, phe duyet, QC, In-Transit, nhan hang',
        'be_keywords': ['TransferRequest', 'Transfer', 'InterWarehouseTransfer'],
        'fe_keywords': ['TransferRequest', 'inter-warehouse-transfer', 'validation'],
        'scenarios': ['Tao yeu cau dieu chuyen', 'Submit & Approve yeu cau', 'Chuyen thanh Transfer',
                      'QC Outbound tai kho nguon', 'Ban giao hang cho tai xe',
                      'Nhan hang tai kho dich', 'QC nhan hang tai kho dich',
                      'Xu ly hang tra ve nguon', 'Chenh lech so luong'],
    },
    {
        'id': '006', 'name': 'Stocktake & Adjustment', 'short': 'STK-006',
        'req': 'Kiem tra quy trinh kiem ke kho, xu ly chenh lech va dieu chinh ton kho',
        'be_keywords': ['StockTake', 'Adjustment'],
        'fe_keywords': ['StocktakeList', 'stocktake'],
        'scenarios': ['Tao phieu kiem ke', 'Nhap so luong thuc dem', 'Submit phieu kiem ke',
                      'Phe duyet chenh lech (Variance)', 'Dieu chinh ton kho (Adjustment)',
                      'Loc & tim kiem phieu kiem ke'],
    },
    {
        'id': '007', 'name': 'Pricing & COGS', 'short': 'PRC-007',
        'req': 'Kiem tra quan ly gia ban, lich su gia va tinh gia von hang ban (COGS theo FIFO)',
        'be_keywords': ['PriceHistory', 'Price'],
        'fe_keywords': ['pricing'],
        'scenarios': ['Tao bang gia (Price History)', 'Phe duyet gia', 'Xem lich su gia', 'Tinh COGS'],
    },
    {
        'id': '008', 'name': 'Finance, Billing & Closing', 'short': 'FIN-008',
        'req': 'Kiem tra quy trinh tai chinh: hoa don, thanh toan, cong no dai ly, dong ky ke toan',
        'be_keywords': ['Invoice', 'Payment', 'AccountingPeriod', 'AutoInvoice', 'SupplierInvoice',
                        'SupplierPayment', 'BillingNotification'],
        'fe_keywords': ['PeriodClosing', 'SupplierInvoices', 'DealerDebtInvoice', 'PaymentReceipts', 'finance'],
        'scenarios': ['Tao hoa don tu dong (Auto Invoice)', 'Hoa don NCC (Supplier Invoice)',
                      'Thanh toan NCC', 'Quan ly cong no dai ly', 'Dong ky ke toan (Period Closing)',
                      'Bao cao cong no (Credit Aging)'],
    },
    {
        'id': '009', 'name': 'Returns, Scrap & Disposal', 'short': 'RET-009',
        'req': 'Kiem tra quy trinh tra hang khach, cap credit note va xu ly hang tieu huy',
        'be_keywords': ['Returns', 'Disposal', 'ReceiptRtv', 'ReceiptServiceReturn'],
        'fe_keywords': ['returns'],
        'scenarios': ['Tao phieu tra hang khach', 'QC hang tra', 'Cap Credit Note', 'Xu ly hang tieu huy'],
    },
    {
        'id': '010', 'name': 'Reports, Dashboard & Alerts', 'short': 'RPT-010',
        'req': 'Kiem tra bao cao, dashboard CEO va canh bao ton kho thap',
        'be_keywords': ['Report', 'StockAlert', 'ReportService'],
        'fe_keywords': ['report', 'CreditAgingReport'],
        'scenarios': ['CEO Dashboard', 'Bao cao dinh gia ton kho', 'Canh bao ton kho thap', 'Bao cao cong no'],
    },
]

# ============================================================
# HELPERS
# ============================================================
def parse_surefire(surefire_dir: Path) -> dict:
    results = {}
    if not surefire_dir.exists():
        return results
    today = datetime.now().strftime('%Y-%m-%d')
    for xf in surefire_dir.glob('TEST-*.xml'):
        try:
            tree = ET.parse(xf)
            root = tree.getroot()
            suite = root.get('name', '')
            ts = root.get('timestamp', today)
            if ts and 'T' in ts:
                ts = ts.split('T')[0]
            for tc in root.findall('testcase'):
                method = tc.get('name', '')
                cls    = tc.get('classname', suite).split('.')[-1]
                if tc.find('failure') is not None or tc.find('error') is not None:
                    st = 'Failed'
                elif tc.find('skipped') is not None:
                    st = 'N/A'
                else:
                    st = 'Passed'
                results.setdefault(cls, {})[method] = {'status': st, 'date': ts}
        except Exception:
            continue
    return results


def snake_to_words(name: str) -> str:
    parts = name.split('_')
    words = []
    for p in parts:
        words.append(re.sub(r'([A-Z])', r' \1', p).strip().lower())
    return ' '.join(words)


def make_procedure(method: str) -> str:
    w = snake_to_words(method).lower()
    if any(x in w for x in ['login', 'auth', 'signin']):
        return '1. Mo trinh duyet, truy cap trang dang nhap\n2. Nhap thong tin tai khoan\n3. Nhan Dang nhap'
    elif any(x in w for x in ['create', 'add']):
        return '1. Dang nhap voi role phu hop\n2. Dieu huong den chuc nang\n3. Nhan Tao moi\n4. Dien thong tin\n5. Nhan Luu / Submit'
    elif any(x in w for x in ['update', 'edit']):
        return '1. Dang nhap voi role phu hop\n2. Mo ban ghi can cap nhat\n3. Nhan Chinh sua\n4. Thay doi thong tin\n5. Nhan Luu'
    elif any(x in w for x in ['delete', 'deactivate', 'disable']):
        return '1. Dang nhap voi role phu hop\n2. Tim ban ghi can xu ly\n3. Nhan Vo hieu hoa / Xoa\n4. Xac nhan thao tac'
    elif any(x in w for x in ['approve', 'phe duyet']):
        return '1. Dang nhap voi role co quyen duyet\n2. Mo danh sach cho duyet\n3. Chon phieu\n4. Nhan Phe duyet'
    elif any(x in w for x in ['reject', 'cancel']):
        return '1. Dang nhap voi role phu hop\n2. Mo phieu can xu ly\n3. Nhan Tu choi / Huy\n4. Nhap ly do'
    elif any(x in w for x in ['list', 'search', 'filter', 'getall']):
        return '1. Dang nhap he thong\n2. Dieu huong den danh sach\n3. Ap dung bo loc\n4. Quan sat ket qua'
    else:
        return f'1. Dang nhap he thong\n2. Thuc hien: {w[:60]}\n3. Quan sat ket qua'


def make_expected(method: str) -> str:
    w = method.lower()
    if any(x in w for x in ['success', 'valid', 'returns200', 'returns201', 'returns204']):
        return 'He thong xu ly thanh cong.\nThong bao success hien thi.\nDu lieu duoc luu dung.'
    elif any(x in w for x in ['rejects', 'fails', 'returns400', 'throws', 'invalid', 'error']):
        return 'He thong hien thi thong bao loi.\nKhong co du lieu bi thay doi.\nHTTP 4xx tra ve.'
    elif any(x in w for x in ['returns401', 'unauthenticated', 'unauthorized']):
        return 'He thong tu choi truy cap. HTTP 401 Unauthorized.'
    elif any(x in w for x in ['returns403', 'forbidden']):
        return 'He thong tu choi quyen. HTTP 403 Forbidden.'
    elif any(x in w for x in ['notfound', 'returns404']):
        return 'He thong tra ve khong tim thay. HTTP 404.'
    else:
        return 'He thong phan hoi dung nghiep vu.\nDu lieu hien thi chinh xac.'


def extract_precond(java_content: str) -> str:
    match = re.search(
        r'@BeforeEach[\s\S]*?void\s+\w+\s*\(\s*\)\s*\{([\s\S]*?)(?=\n\s*@|\Z)',
        java_content)
    conds = []
    if match:
        body = match.group(1)
        if re.search(r'new\s+User|mockUser|adminUser|managerUser', body):
            conds.append('Tai khoan nguoi dung da duoc tao')
        if re.search(r'new\s+Warehouse|mockWarehouse|warehouse\s*=', body):
            conds.append('Kho hang da duoc khoi tao')
        if re.search(r'new\s+Product|mockProduct|product\s*=', body):
            conds.append('San pham da duoc dang ky')
        if re.search(r'JWT|token|jwtToken|authHeader', body):
            conds.append('Token xac thuc hop le da duoc cap phat')
        if re.search(r'new\s+Receipt|receipt\s*=', body):
            conds.append('Phieu nhap kho da ton tai')
        if re.search(r'new\s+Trip|trip\s*=', body):
            conds.append('Chuyen hang da duoc khoi tao')
    if not conds:
        conds = ['He thong dang hoat dong binh thuong', 'Nguoi dung da dang nhap voi role phu hop']
    return '\n'.join(f'- {c}' for c in conds)


def get_tcs_for_module(module: dict, surefire: dict) -> list:
    """Returns list of (tc_id, description, procedure, expected, precond, r1_status, r1_date, r1_tester, note)"""
    rows = []
    scenario_counter = {}  # scenario_name -> count

    # ---- Backend ----
    for root, _, files in os.walk(BE_TEST_DIR):
        for f in files:
            if not f.endswith('.java'):
                continue
            cls = f.replace('.java', '')
            if not any(kw.lower() in cls.lower() for kw in module['be_keywords']):
                continue
            path = Path(root) / f
            content = path.read_text(encoding='utf-8', errors='ignore')
            precond = extract_precond(content)
            methods = re.findall(r'@Test[\s\S]{0,300}?void\s+(\w+)\s*\(', content)
            cls_results = surefire.get(cls, {})
            for i, method in enumerate(methods):
                res = cls_results.get(method, {})
                st  = res.get('status', 'Pending')
                dt  = res.get('date', '')
                tester = TESTER_AUTO if st in ('Passed', 'Failed') else ''
                # Assign scenario letter
                scenario_idx = (len(rows) // 5) % len(module['scenarios'])
                scenario = module['scenarios'][scenario_idx]
                rows.append({
                    'scenario': scenario,
                    'id':  f"{module['short']}-{len(rows)+1:03d}",
                    'desc': snake_to_words(method).title(),
                    'proc': make_procedure(method),
                    'exp':  make_expected(method),
                    'pre':  precond,
                    'r1':   st,
                    'dt':   dt,
                    'tester': tester,
                    'note': f'Source: {cls}.java',
                })

    # ---- Frontend ----
    pattern = re.compile(r"""(?:it|test)\s*\(['"](.*?)['"]\s*,""")
    for base in FE_TEST_DIRS:
        if not base.exists():
            continue
        for root, _, files in os.walk(base):
            for f in files:
                if '.test.' not in f:
                    continue
                if not any(kw.lower() in f.lower() for kw in module['fe_keywords']):
                    continue
                path = Path(root) / f
                content = path.read_text(encoding='utf-8', errors='ignore')
                for m in pattern.findall(content):
                    scenario_idx = (len(rows) // 5) % len(module['scenarios'])
                    scenario = module['scenarios'][scenario_idx]
                    rows.append({
                        'scenario': scenario,
                        'id': f"{module['short']}-FE{len(rows)+1:03d}",
                        'desc': m,
                        'proc': make_procedure(m),
                        'exp':  make_expected(m),
                        'pre':  '- Trinh duyet da mo ung dung WMS\n- Nguoi dung da dang nhap',
                        'r1':   'Pending',
                        'dt':   '',
                        'tester': '',
                        'note': f'Frontend: {f}',
                    })
    return rows


# ============================================================
# COPY CELL STYLE (minimal — copy fill/font/alignment/border)
# ============================================================
def copy_cell_style(src_cell, dst_cell):
    """Copy basic style from src to dst, preserving template look."""
    if src_cell.has_style:
        if src_cell.font:
            dst_cell.font = copy.copy(src_cell.font)
        if src_cell.fill:
            dst_cell.fill = copy.copy(src_cell.fill)
        if src_cell.border:
            dst_cell.border = copy.copy(src_cell.border)
        if src_cell.alignment:
            dst_cell.alignment = copy.copy(src_cell.alignment)
        if src_cell.number_format:
            dst_cell.number_format = src_cell.number_format


# ============================================================
# CLONE WORKSHEET (copy sheet structure + formatting from template)
# ============================================================
def clone_workflow_sheet(wb, src_sheet_name: str, new_name: str) -> Worksheet:
    """Clone a workflow sheet from the workbook and rename it."""
    src = wb[src_sheet_name]
    tgt = wb.copy_worksheet(src)
    tgt.title = new_name[:31]
    return tgt


# ============================================================
# FILL WORKFLOW SHEET WITH DATA
# ============================================================
STATUS_FILL = {
    'Passed':  PatternFill('solid', fgColor='C6EFCE'),
    'Failed':  PatternFill('solid', fgColor='FFC7CE'),
    'Pending': PatternFill('solid', fgColor='FFEB9C'),
    'N/A':     PatternFill('solid', fgColor='D9D9D9'),
}
SCENARIO_FILL = PatternFill('solid', fgColor='BDD7EE')  # light blue
SCENARIO_FONT = Font(name='Calibri', bold=True, size=10, color='1F3864')
DATA_ALIGN    = Alignment(wrap_text=True, vertical='top')
STATUS_ALIGN  = Alignment(horizontal='center', vertical='center', wrap_text=False)

def fill_module_sheet(ws: Worksheet, module: dict, tcs: list):
    """Fill module name, requirement, and all TC rows into the cloned sheet."""
    # Row 1: Workflow name (col B)
    ws['B1'] = module['name']
    # Row 2: Test requirement (col B)
    ws['B2'] = module['req']
    # Row 3: Number of TCs = COUNTA(A12:A5000) — keep formula
    ws['B3'] = '=COUNTA(A12:A5000)'
    # Row 5: Round summaries — use COUNTIF on col F (Round 1)
    # Rows 5,6,7 = Round 1,2,3 with B=Passed, C=Failed, D=Pending, E=N/A
    for rnd_row in [5, 6, 7]:
        ws.cell(row=rnd_row, column=2).value = '=COUNTIF($F$12:$F$5000,B4)'
        ws.cell(row=rnd_row, column=3).value = '=COUNTIF($F$12:$F$5000,C4)'
        ws.cell(row=rnd_row, column=4).value = '=COUNTIF($F$12:$F$5000,D4)'
        ws.cell(row=rnd_row, column=5).value = '=COUNTIF($F$12:$F$5000,E4)'

    # Clear template example rows (rows 10-end)
    # Template has row 9 as header, row 10=Scenario A, rows 11-onwards = example data
    # We keep row 9 (header). Clear from row 10 downward
    max_row_to_clear = max(ws.max_row, 30)
    for r in range(10, max_row_to_clear + 1):
        for c in range(1, 16):
            ws.cell(row=r, column=c).value = None

    # Now fill actual data starting row 10
    current_row = 10
    current_scenario = None

    for tc in tcs:
        # Insert scenario separator when scenario changes
        if tc['scenario'] != current_scenario:
            current_scenario = tc['scenario']
            ws.cell(row=current_row, column=1).value = current_scenario
            ws.cell(row=current_row, column=1).font = SCENARIO_FONT
            ws.cell(row=current_row, column=1).fill = SCENARIO_FILL
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
            # Merge across all 15 cols for visual clarity (optional — minimal change)
            # Actually keep simple to avoid merge conflicts
            current_row += 1

        # TC row
        r = current_row
        ws.cell(row=r, column=1).value  = tc['id']
        ws.cell(row=r, column=2).value  = tc['desc']
        ws.cell(row=r, column=3).value  = tc['proc']
        ws.cell(row=r, column=4).value  = tc['exp']
        ws.cell(row=r, column=5).value  = tc['pre']
        # Round 1
        ws.cell(row=r, column=6).value  = tc['r1']
        ws.cell(row=r, column=7).value  = tc['dt']
        ws.cell(row=r, column=8).value  = tc['tester']
        # Round 2 & 3 — leave Pending
        ws.cell(row=r, column=9).value  = 'Pending'
        ws.cell(row=r, column=12).value = 'Pending'

        # Apply wrap alignment to text cols
        for c in [1, 2, 3, 4, 5, 8, 15]:
            ws.cell(row=r, column=c).alignment = DATA_ALIGN

        # Status color
        r1_val = tc['r1']
        if r1_val in STATUS_FILL:
            ws.cell(row=r, column=6).fill = STATUS_FILL[r1_val]
        ws.cell(row=r, column=6).alignment  = STATUS_ALIGN
        ws.cell(row=r, column=9).fill = STATUS_FILL['Pending']
        ws.cell(row=r, column=9).alignment  = STATUS_ALIGN
        ws.cell(row=r, column=12).fill = STATUS_FILL['Pending']
        ws.cell(row=r, column=12).alignment = STATUS_ALIGN

        ws.row_dimensions[r].height = 55
        current_row += 1

    return current_row  # next available row


# ============================================================
# UPDATE TEST STATISTICS SHEET
# ============================================================
def update_statistics_sheet(wb, module_sheet_names: list):
    ws = wb['Test Statistics']

    # Fill project info
    ws['C3'] = PROJECT_NAME
    ws['C4'] = PROJECT_CODE
    ws['H5'] = RUN_DATE
    ws['C6'] = f'Sprint 1 - Release 1.0 | {len(module_sheet_names)} modules: ' + ', '.join(
        m['short'] for m in MODULES)

    # Clear existing module rows (rows 11-onward up to summary)
    for r in range(11, 25):
        for c in range(2, 9):
            ws.cell(row=r, column=c).value = None

    # Write module rows
    for i, (mod, sname) in enumerate(zip(MODULES, module_sheet_names)):
        row = 11 + i
        # Escape sheet name for formula
        safe = f"'{sname}'" if ' ' in sname or '&' in sname else sname
        ws.cell(row=row, column=2).value = i + 1
        ws.cell(row=row, column=3).value = f'={safe}!B1'
        ws.cell(row=row, column=4).value = f'={safe}!B5'   # Round 1 Passed
        ws.cell(row=row, column=5).value = f'={safe}!C5'   # Round 1 Failed
        ws.cell(row=row, column=6).value = f'={safe}!D5'   # Round 1 Pending
        ws.cell(row=row, column=7).value = f'={safe}!E5'   # Round 1 N/A
        ws.cell(row=row, column=8).value = f'={safe}!B3'   # Number of TCs

    # Sub total row
    n = len(MODULES)
    sub_row = 11 + n
    ws.cell(row=sub_row, column=3).value = 'Sub total'
    ws.cell(row=sub_row, column=4).value = f'=SUM(D11:D{sub_row-1})'
    ws.cell(row=sub_row, column=5).value = f'=SUM(E11:E{sub_row-1})'
    ws.cell(row=sub_row, column=6).value = f'=SUM(F11:F{sub_row-1})'
    ws.cell(row=sub_row, column=7).value = f'=SUM(G11:G{sub_row-1})'
    ws.cell(row=sub_row, column=8).value = f'=SUM(H11:H{sub_row-1})'
    # Coverage
    ws.cell(row=sub_row+2, column=3).value = 'Test coverage'
    ws.cell(row=sub_row+2, column=5).value = f'=(D{sub_row}+E{sub_row})*100/(H{sub_row}-G{sub_row})'
    ws.cell(row=sub_row+2, column=6).value = '%'
    ws.cell(row=sub_row+3, column=3).value = 'Test successful coverage'
    ws.cell(row=sub_row+3, column=5).value = f'=D{sub_row}*100/(H{sub_row}-G{sub_row})'
    ws.cell(row=sub_row+3, column=6).value = '%'


# ============================================================
# UPDATE TEST CASES SHEET
# ============================================================
def update_test_cases_sheet(wb, module_sheet_names: list):
    ws = wb['Test Cases']
    # Project info
    ws['D3'] = PROJECT_NAME
    ws['D4'] = PROJECT_CODE
    ws['D5'] = (
        'He thong WMS chay tren:\n'
        '1. Backend: Spring Boot 3.4.5 / Java 21 / PostgreSQL 18\n'
        '2. Frontend: React 18\n'
        '3. Moi truong test: localhost / staging server\n'
        '4. Browser: Chrome latest'
    )

    # Clear old function rows
    for r in range(9, 30):
        for c in range(2, 7):
            ws.cell(row=r, column=c).value = None

    # Fill function list
    for i, (mod, sname) in enumerate(zip(MODULES, module_sheet_names)):
        r = 9 + i
        ws.cell(row=r, column=2).value = i + 1
        ws.cell(row=r, column=3).value = mod['name']
        ws.cell(row=r, column=4).value = sname
        ws.cell(row=r, column=5).value = mod['req']


# ============================================================
# UPDATE COVER SHEET
# ============================================================
def update_cover_sheet(wb):
    ws = wb['Cover']
    ws['B4'] = PROJECT_NAME
    ws['B5'] = PROJECT_CODE
    ws['B6'] = '=B5&"_"&"SystemTest"&"_"&"v1.0"'
    ws['F5'] = RUN_DATE


# ============================================================
# MAIN
# ============================================================
def main():
    print('=' * 60)
    print('WMS System Test — Fill Template')
    print(f'Run date: {RUN_DATE}')
    print('=' * 60)

    # 1. Parse surefire
    print('\n[1/5] Parsing Maven Surefire reports...')
    surefire = parse_surefire(SUREFIRE)
    print(f'  -> {len(surefire)} test classes found')

    # 2. Load template
    print('\n[2/5] Loading template...')
    wb = load_workbook(str(TEMPLATE))
    print(f'  -> Sheets: {wb.sheetnames}')

    # We need: Cover, Test Cases, Test Statistics, Workflow Name1, Workflow Name2
    # Strategy:
    # - Rename "Workflow Name1" -> first module
    # - Clone "Workflow Name2" (8 times) for modules 3-10
    # - Rename "Workflow Name2" -> second module

    # 3. Rename existing workflow sheets
    print('\n[3/5] Setting up module sheets...')
    module_sheet_names = []

    # Rename Workflow Name1 -> module 0
    ws1 = wb['Workflow Name1']
    ws1.title = MODULES[0]['short'][:31]
    module_sheet_names.append(ws1.title)
    print(f'  Renamed "Workflow Name1" -> {ws1.title}')

    # Rename Workflow Name2 -> module 1
    ws2 = wb['Workflow Name2']
    ws2.title = MODULES[1]['short'][:31]
    module_sheet_names.append(ws2.title)
    print(f'  Renamed "Workflow Name2" -> {ws2.title}')

    # Clone Workflow Name2 (now renamed) for modules 2-9
    # We need to clone from the renamed sheet
    for mod in MODULES[2:]:
        new_title = mod['short'][:31]
        # wb.copy_worksheet copies to end
        new_ws = wb.copy_worksheet(wb[MODULES[1]['short'][:31]])
        new_ws.title = new_title
        module_sheet_names.append(new_title)
        print(f'  Cloned -> {new_title}')

    print(f'  Total module sheets: {len(module_sheet_names)}')

    # 4. Fill data into each module sheet
    print('\n[4/5] Filling test case data...')
    for mod, sname in zip(MODULES, module_sheet_names):
        print(f'  [{mod["id"]}] {mod["name"]}...')
        tcs = get_tcs_for_module(mod, surefire)
        ws = wb[sname]
        fill_module_sheet(ws, mod, tcs)
        passed = sum(1 for t in tcs if t['r1'] == 'Passed')
        failed = sum(1 for t in tcs if t['r1'] == 'Failed')
        print(f'      {len(tcs)} TCs | Passed={passed} Failed={failed} Pending={len(tcs)-passed-failed}')

    # 5. Update meta sheets
    print('\n[5/5] Updating meta sheets...')
    update_statistics_sheet(wb, module_sheet_names)
    update_test_cases_sheet(wb, module_sheet_names)
    update_cover_sheet(wb)

    # Save
    out_name = f'SystemTest_WMS_{RUN_DATE.replace("-", "")}.xlsx'
    out_path = OUT_DIR / out_name
    wb.save(str(out_path))
    print(f'\nSaved: {out_path}')
    print(f'Sheets: {wb.sheetnames}')
    return str(out_path)


if __name__ == '__main__':
    main()
