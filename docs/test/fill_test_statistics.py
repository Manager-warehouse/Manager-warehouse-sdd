# -*- coding: utf-8 -*-
"""
Script to completely fill Test Statistics sheet with metadata, formulas, hyperlinks, formatting, and colors
"""

import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(r'd:\swp\Manager-warehouse-sdd')
EXCEL_OUT = ROOT / 'docs' / 'test' / 'test_final.xlsx'

MODULE_CODES = [
    ("AUTH-001", "Security, Auth & RBAC"),
    ("MDM-002", "Master Data Management"),
    ("RCV-003", "Inbound Receipt & QC"),
    ("OUT-004", "Outbound Delivery & POD"),
    ("TRF-005", "Inter-Warehouse Transfer"),
    ("STK-006", "Stocktake & Adjustment"),
    ("PRC-007", "Pricing & COGS"),
    ("FIN-008", "Finance, Billing & Closing"),
    ("RET-009", "Returns, Scrap & Disposal"),
    ("RPT-010", "Reports, Dashboard & Alerts"),
]

def fill_test_statistics():
    wb = openpyxl.load_workbook(str(EXCEL_OUT))
    ws = wb['Test Statistics']

    tahoma_regular = Font(name='Tahoma', size=10, bold=False, color='000000')
    tahoma_bold = Font(name='Tahoma', size=10, bold=True, color='000000')
    tahoma_link = Font(name='Tahoma', size=10, bold=False, color='0000FF', underline='single')

    align_left = Alignment(horizontal='left', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )

    no_fill = PatternFill(fill_type=None)
    subtotal_fill = PatternFill('solid', fgColor='D9E1F2')

    # 1. Fill Header Metadata (Rows 3 to 6)
    ws['C3'] = 'Warehouse Management System (WMS)'
    ws['C3'].font = tahoma_regular

    ws['C4'] = 'WMS-SPRINT1'
    ws['C4'].font = tahoma_regular

    ws['C5'] = '=C4&"_"&"Test_Report"&"_"&"v1.0"'
    ws['C5'].font = tahoma_regular

    ws['C6'] = 'Sprint 1 - Release 1.0 | 10 modules: AUTH-001, MDM-002, RCV-003, OUT-004, TRF-005, STK-006, PRC-007, FIN-008, RET-009, RPT-010'
    ws['C6'].font = tahoma_regular

    # Creator (G3)
    ws['G3'] = 'QA Team / Dev Engineering'
    ws['G3'].font = tahoma_regular

    # Reviewer (G4)
    ws['G4'] = 'Project Lead / Director'
    ws['G4'].font = tahoma_regular

    # Issue Date (H5)
    ws['H5'] = '2026-07-23'
    ws['H5'].font = tahoma_regular
    ws['H5'].alignment = align_center

    # 2. Reset rows 11 to 25 styles & clear old dark fills
    for r in range(11, 26):
        ws.row_dimensions[r].height = 22
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = no_fill
            cell.hyperlink = None

    # 3. Fill Module Rows 11 to 20
    for i, (code, name) in enumerate(MODULE_CODES, start=1):
        r = 10 + i
        safe_code = f"'{code}'"

        # Col B (No)
        c_no = ws.cell(row=r, column=2)
        c_no.value = i
        c_no.font = tahoma_regular
        c_no.alignment = align_center
        c_no.border = thin_border

        # Col C (Module code) -> Clickable Hyperlink pointing to sheet
        c_mod = ws.cell(row=r, column=3)
        c_mod.value = f'=HYPERLINK("#\'{code}\'!A1", {safe_code}!B1)'
        c_mod.font = tahoma_link
        c_mod.alignment = align_left
        c_mod.border = thin_border
        c_mod.hyperlink = Hyperlink(ref=f'C{r}', location=f"'{code}'!A1", display=name)

        # Col D (Passed)
        c_p = ws.cell(row=r, column=4)
        c_p.value = f'={safe_code}!B5'
        c_p.font = tahoma_regular
        c_p.alignment = align_center
        c_p.border = thin_border

        # Col E (Failed)
        c_f = ws.cell(row=r, column=5)
        c_f.value = f'={safe_code}!C5'
        c_f.font = tahoma_regular
        c_f.alignment = align_center
        c_f.border = thin_border

        # Col F (Pending)
        c_pend = ws.cell(row=r, column=6)
        c_pend.value = f'={safe_code}!D5'
        c_pend.font = tahoma_regular
        c_pend.alignment = align_center
        c_pend.border = thin_border

        # Col G (N/A)
        c_na = ws.cell(row=r, column=7)
        c_na.value = f'={safe_code}!E5'
        c_na.font = tahoma_regular
        c_na.alignment = align_center
        c_na.border = thin_border

        # Col H (Number of TCs)
        c_tot = ws.cell(row=r, column=8)
        c_tot.value = f'={safe_code}!B3'
        c_tot.font = tahoma_regular
        c_tot.alignment = align_center
        c_tot.border = thin_border

    # 4. Row 21: Sub total
    r_sub = 21
    c_sub_lbl = ws.cell(row=r_sub, column=3)
    c_sub_lbl.value = 'Sub total'
    c_sub_lbl.font = tahoma_bold
    c_sub_lbl.alignment = align_left
    c_sub_lbl.border = thin_border
    c_sub_lbl.fill = subtotal_fill

    for col_idx, col_let in [(4, 'D'), (5, 'E'), (6, 'F'), (7, 'G'), (8, 'H')]:
        c_cell = ws.cell(row=r_sub, column=col_idx)
        c_cell.value = f'=SUM({col_let}11:{col_let}20)'
        c_cell.font = tahoma_bold
        c_cell.alignment = align_center
        c_cell.border = thin_border
        c_cell.fill = subtotal_fill

    ws.cell(row=r_sub, column=2).border = thin_border
    ws.cell(row=r_sub, column=2).fill = subtotal_fill

    # 5. Row 23: Test coverage
    r_cov = 23
    c_cov_lbl = ws.cell(row=r_cov, column=3)
    c_cov_lbl.value = 'Test coverage'
    c_cov_lbl.font = tahoma_bold
    c_cov_lbl.alignment = align_left

    c_cov_val = ws.cell(row=r_cov, column=5)
    c_cov_val.value = '=(D21+E21)*100/(H21-G21)'
    c_cov_val.font = tahoma_bold
    c_cov_val.alignment = align_right

    c_cov_unit = ws.cell(row=r_cov, column=6)
    c_cov_unit.value = '%'
    c_cov_unit.font = tahoma_regular
    c_cov_unit.alignment = align_left

    # 6. Row 24: Test successful coverage
    r_pass = 24
    c_pass_lbl = ws.cell(row=r_pass, column=3)
    c_pass_lbl.value = 'Test successful coverage'
    c_pass_lbl.font = tahoma_bold
    c_pass_lbl.alignment = align_left

    c_pass_val = ws.cell(row=r_pass, column=5)
    c_pass_val.value = '=D21*100/(H21-G21)'
    c_pass_val.font = tahoma_bold
    c_pass_val.alignment = align_right

    c_pass_unit = ws.cell(row=r_pass, column=6)
    c_pass_unit.value = '%'
    c_pass_unit.font = tahoma_regular
    c_pass_unit.alignment = align_left

    wb.save(str(EXCEL_OUT))
    print(f"Successfully filled Test Statistics sheet in {EXCEL_OUT}")

if __name__ == '__main__':
    fill_test_statistics()
