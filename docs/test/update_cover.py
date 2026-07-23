# -*- coding: utf-8 -*-
"""
Script to update Cover sheet in test_final.xlsx with complete metadata
"""

import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(r'd:\swp\Manager-warehouse-sdd')
EXCEL_OUT = ROOT / 'docs' / 'test' / 'test_final.xlsx'

def update_cover():
    wb = openpyxl.load_workbook(str(EXCEL_OUT))
    ws = wb['Cover']

    tahoma_regular = Font(name='Tahoma', size=10, bold=False)
    tahoma_bold = Font(name='Tahoma', size=10, bold=True)
    align_left = Alignment(horizontal='left', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')
    align_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )

    # 1. Fill Metadata Header
    ws['B4'] = 'Warehouse Management System (WMS)'
    ws['B4'].font = tahoma_regular

    ws['B5'] = 'WMS-SPRINT1'
    ws['B5'].font = tahoma_regular

    ws['B6'] = '=B5&"_"&"SystemTest"&"_"&"v1.0"'
    ws['B6'].font = tahoma_regular

    # Creator (F4)
    ws['F4'] = 'QA Team / Dev Engineering'
    ws['F4'].font = tahoma_regular
    ws['F4'].alignment = align_left

    # Issue Date (F5)
    ws['F5'] = '2026-07-23'
    ws['F5'].font = tahoma_regular
    ws['F5'].alignment = align_left

    # Version (F6)
    ws['F6'] = 'v1.0'
    ws['F6'].font = tahoma_regular
    ws['F6'].alignment = align_left

    # 2. Record of Change Table (Row 11)
    ws['A11'] = '2026-07-23'
    ws['A11'].font = tahoma_regular
    ws['A11'].alignment = align_center

    ws['B11'] = 'v1.0'
    ws['B11'].font = tahoma_regular
    ws['B11'].alignment = align_center

    ws['C11'] = 'System Test Report Document'
    ws['C11'].font = tahoma_regular
    ws['C11'].alignment = align_left

    ws['D11'] = 'A'
    ws['D11'].font = tahoma_regular
    ws['D11'].alignment = align_center

    ws['E11'] = 'Khởi tạo báo cáo System Test toàn bộ 10 module Sprint 1 (924 Test Cases)'
    ws['E11'].font = tahoma_regular
    ws['E11'].alignment = align_wrap

    ws['F11'] = 'CLAUDE.md; .sdd/specs/001-010; result_test.md'
    ws['F11'].font = tahoma_regular
    ws['F11'].alignment = align_wrap

    # Apply borders for Row 11
    for col_idx in range(1, 7):
        ws.cell(row=11, column=col_idx).border = thin_border

    ws.row_dimensions[11].height = 30

    wb.save(str(EXCEL_OUT))
    print(f"Successfully updated Cover sheet in {EXCEL_OUT}")

if __name__ == '__main__':
    update_cover()
