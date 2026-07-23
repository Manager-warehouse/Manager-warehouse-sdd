# -*- coding: utf-8 -*-
"""
Script to fix sheet hyperlinks according to exact user feedback:
1. Function Name (Col C) -> Plain text (NO hyperlink formula, NO hyperlink object, Tahoma 10pt black text)
2. Sheet Name (Col D) -> Native + Formula hyperlink pointing to the EXACT sheet (e.g. 'AUTH-001'!A1, 'MDM-002'!A1)
"""

import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(r'd:\swp\Manager-warehouse-sdd')
EXCEL_OUT = ROOT / 'docs' / 'test' / 'test_final.xlsx'

MODULE_INFO = [
    (1, "Security, Auth & RBAC", "AUTH-001"),
    (2, "Master Data Management", "MDM-002"),
    (3, "Inbound Receipt & QC", "RCV-003"),
    (4, "Outbound Delivery & POD", "OUT-004"),
    (5, "Inter-Warehouse Transfer", "TRF-005"),
    (6, "Stocktake & Adjustment", "STK-006"),
    (7, "Pricing & COGS", "PRC-007"),
    (8, "Finance, Billing & Closing", "FIN-008"),
    (9, "Returns, Scrap & Disposal", "RET-009"),
    (10, "Reports, Dashboard & Alerts", "RPT-010"),
]

def fix_hyperlinks():
    wb = openpyxl.load_workbook(str(EXCEL_OUT))

    tahoma_regular = Font(name='Tahoma', size=10, bold=False, color='000000')
    tahoma_link = Font(name='Tahoma', size=10, bold=False, color='0000FF', underline='single')

    align_left = Alignment(horizontal='left', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )

    ws_tc = wb['Test Cases']

    # Clear old rows 9 to 30 for clean rebuild
    for r in range(9, 35):
        for c in range(2, 7):
            cell = ws_tc.cell(row=r, column=c)
            cell.value = None
            cell.hyperlink = None

    for i, (no, func_name, code) in enumerate(MODULE_INFO, start=1):
        row = 8 + i

        # Col B (No)
        c_no = ws_tc.cell(row=row, column=2)
        c_no.value = no
        c_no.font = tahoma_regular
        c_no.alignment = align_center
        c_no.border = thin_border

        # Col C (Function Name) -> PLAIN TEXT (No hyperlink)
        c_func = ws_tc.cell(row=row, column=3)
        c_func.value = func_name
        c_func.font = tahoma_regular
        c_func.alignment = align_left
        c_func.border = thin_border
        c_func.hyperlink = None  # Remove any hyperlink object

        # Col D (Sheet Name) -> ONLY HYPERLINK
        c_sheet = ws_tc.cell(row=row, column=4)
        c_sheet.value = code
        c_sheet.font = tahoma_link
        c_sheet.alignment = align_center
        c_sheet.border = thin_border
        
        # Set BOTH native openpyxl hyperlink location AND formula if needed
        c_sheet.hyperlink = Hyperlink(ref=f'D{row}', location=f"'{code}'!A1", display=code)

        # Col E (Description)
        c_desc = ws_tc.cell(row=row, column=5)
        req_desc = f"Kiểm tra toàn bộ nghiệp vụ thuộc module {func_name}"
        c_desc.value = req_desc
        c_desc.font = tahoma_regular
        c_desc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c_desc.border = thin_border

        ws_tc.row_dimensions[row].height = 25

    wb.save(str(EXCEL_OUT))
    print(f"Successfully fixed hyperlinks in {EXCEL_OUT}")

if __name__ == '__main__':
    fix_hyperlinks()
