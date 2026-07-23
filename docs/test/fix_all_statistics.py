# -*- coding: utf-8 -*-
"""
Script to fix all statistics formulas and header structure across all module sheets and Test Statistics sheet
"""

import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(r'd:\swp\Manager-warehouse-sdd')
EXCEL_OUT = ROOT / 'docs' / 'test' / 'test_final.xlsx'

MODULE_CODES = [
    ("AUTH-001", "Security, Auth & RBAC"),
    ("MDM-002", "Master Data Management"),
    ("RCV-003", "Inbound Receipt & QC"),
    ("OUT-004", "Outbound Delivery & POD"),
    ("TRF-005", "Inter-Warehouse Transfer"),
    ("STK-006", "Stocktake & Adjustment"),
    ("PRC-007", "Pricing & COGS"),
    ("FIN-008", "Finance, Billing & Closing"),
    ("RET-009", "Returns, Scrap & Disposal"),
    ("RPT-010", "Reports, Dashboard & Alerts"),
]

def fix_statistics():
    wb = openpyxl.load_workbook(str(EXCEL_OUT))

    tahoma_regular = Font(name='Tahoma', size=10, bold=False, color='000000')
    tahoma_bold = Font(name='Tahoma', size=10, bold=True, color='000000')
    tahoma_header = Font(name='Tahoma', size=10, bold=True, color='FFFFFF')
    tahoma_link = Font(name='Tahoma', size=10, bold=False, color='0000FF', underline='single')

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    header_fill = PatternFill('solid', fgColor='1F497D')
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )

    # 1. Fix each Module Sheet (Header rows 4 to 8)
    for code, name in MODULE_CODES:
        ws = wb[code]

        # Row 4: Number of TCs
        ws['A4'] = 'Number of TCs'
        ws['A4'].font = tahoma_bold
        ws['B4'] = '=COUNTA(A11:A5000)'
        ws['B4'].font = tahoma_bold

        # Row 5: Testing Round headers
        ws['A5'] = 'Testing Round'
        ws['A5'].font = tahoma_bold
        
        ws['B5'] = 'Passed'
        ws['B5'].font = tahoma_header
        ws['B5'].fill = header_fill
        ws['B5'].alignment = align_center

        ws['C5'] = 'Failed'
        ws['C5'].font = tahoma_header
        ws['C5'].fill = header_fill
        ws['C5'].alignment = align_center

        ws['D5'] = 'Pending'
        ws['D5'].font = tahoma_header
        ws['D5'].fill = header_fill
        ws['D5'].alignment = align_center

        ws['E5'] = 'N/A'
        ws['E5'].font = tahoma_header
        ws['E5'].fill = header_fill
        ws['E5'].alignment = align_center

        # Row 6: Round 1 formulas
        ws['A6'] = 'Round 1'
        ws['A6'].font = tahoma_bold
        ws['B6'] = '=COUNTIF($F$11:$F$5000, "Passed")'
        ws['B6'].font = tahoma_regular
        ws['C6'] = '=COUNTIF($F$11:$F$5000, "Failed")'
        ws['C6'].font = tahoma_regular
        ws['D6'] = '=COUNTIF($F$11:$F$5000, "Pending")'
        ws['D6'].font = tahoma_regular
        ws['E6'] = '=COUNTIF($F$11:$F$5000, "N/A")'
        ws['E6'].font = tahoma_regular

        # Row 7: Round 2 formulas
        ws['A7'] = 'Round 2'
        ws['A7'].font = tahoma_bold
        ws['B7'] = '=COUNTIF($I$11:$I$5000, "Passed")'
        ws['B7'].font = tahoma_regular
        ws['C7'] = '=COUNTIF($I$11:$I$5000, "Failed")'
        ws['C7'].font = tahoma_regular
        ws['D7'] = '=COUNTIF($I$11:$I$5000, "Pending")'
        ws['D7'].font = tahoma_regular
        ws['E7'] = '=COUNTIF($I$11:$I$5000, "N/A")'
        ws['E7'].font = tahoma_regular

        # Row 8: Round 3 formulas
        ws['A8'] = 'Round 3'
        ws['A8'].font = tahoma_bold
        ws['B8'] = '=COUNTIF($L$11:$L$5000, "Passed")'
        ws['B8'].font = tahoma_regular
        ws['C8'] = '=COUNTIF($L$11:$L$5000, "Failed")'
        ws['C8'].font = tahoma_regular
        ws['D8'] = '=COUNTIF($L$11:$L$5000, "Pending")'
        ws['D8'].font = tahoma_regular
        ws['E8'] = '=COUNTIF($L$11:$L$5000, "N/A")'
        ws['E8'].font = tahoma_regular

    # 2. Fix Test Statistics Sheet
    ws_stat = wb['Test Statistics']
    subtotal_fill = PatternFill('solid', fgColor='D9E1F2')

    for i, (code, name) in enumerate(MODULE_CODES, start=1):
        r = 10 + i
        safe_code = f"'{code}'"

        # Col B (No)
        c_no = ws_stat.cell(row=r, column=2)
        c_no.value = i
        c_no.font = tahoma_regular
        c_no.alignment = align_center
        c_no.border = thin_border

        # Col C (Module code) -> Clickable Link
        c_mod = ws_stat.cell(row=r, column=3)
        c_mod.value = f'=HYPERLINK("#\'{code}\'!A1", {safe_code}!B1)'
        c_mod.font = tahoma_link
        c_mod.alignment = align_left
        c_mod.border = thin_border
        c_mod.hyperlink = Hyperlink(ref=f'C{r}', location=f"'{code}'!A1", display=name)

        # Col D (Passed) -> Points to Round 1 Passed count cell B6 in module sheet
        c_p = ws_stat.cell(row=r, column=4)
        c_p.value = f'={safe_code}!B6'
        c_p.font = tahoma_regular
        c_p.alignment = align_center
        c_p.border = thin_border
        c_p.number_format = '0'

        # Col E (Failed) -> Points to Cell C6
        c_f = ws_stat.cell(row=r, column=5)
        c_f.value = f'={safe_code}!C6'
        c_f.font = tahoma_regular
        c_f.alignment = align_center
        c_f.border = thin_border
        c_f.number_format = '0'

        # Col F (Pending) -> Points to Cell D6
        c_pend = ws_stat.cell(row=r, column=6)
        c_pend.value = f'={safe_code}!D6'
        c_pend.font = tahoma_regular
        c_pend.alignment = align_center
        c_pend.border = thin_border
        c_pend.number_format = '0'

        # Col G (N/A) -> Points to Cell E6
        c_na = ws_stat.cell(row=r, column=7)
        c_na.value = f'={safe_code}!E6'
        c_na.font = tahoma_regular
        c_na.alignment = align_center
        c_na.border = thin_border
        c_na.number_format = '0'

        # Col H (Number of TCs) -> Points to Cell B4
        c_tot = ws_stat.cell(row=r, column=8)
        c_tot.value = f'={safe_code}!B4'
        c_tot.font = tahoma_regular
        c_tot.alignment = align_center
        c_tot.border = thin_border
        c_tot.number_format = '0'

    # Row 21: Sub total
    r_sub = 21
    c_sub_lbl = ws_stat.cell(row=r_sub, column=3)
    c_sub_lbl.value = 'Sub total'
    c_sub_lbl.font = tahoma_bold
    c_sub_lbl.alignment = align_left
    c_sub_lbl.border = thin_border
    c_sub_lbl.fill = subtotal_fill

    for col_idx, col_let in [(4, 'D'), (5, 'E'), (6, 'F'), (7, 'G'), (8, 'H')]:
        c_cell = ws_stat.cell(row=r_sub, column=col_idx)
        c_cell.value = f'=SUM({col_let}11:{col_let}20)'
        c_cell.font = tahoma_bold
        c_cell.alignment = align_center
        c_cell.border = thin_border
        c_cell.fill = subtotal_fill
        c_cell.number_format = '0'

    ws_stat.cell(row=r_sub, column=2).border = thin_border
    ws_stat.cell(row=r_sub, column=2).fill = subtotal_fill

    # Row 23: Test coverage
    r_cov = 23
    c_cov_val = ws_stat.cell(row=r_cov, column=5)
    c_cov_val.value = '=(D21+E21)*100/(H21-G21)'
    c_cov_val.font = tahoma_bold
    c_cov_val.alignment = align_right
    c_cov_val.number_format = '0.0'

    # Row 24: Test successful coverage
    r_pass = 24
    c_pass_val = ws_stat.cell(row=r_pass, column=5)
    c_pass_val.value = '=D21*100/(H21-G21)'
    c_pass_val.font = tahoma_bold
    c_pass_val.alignment = align_right
    c_pass_val.number_format = '0.0'

    wb.save(str(EXCEL_OUT))
    print(f"Successfully fixed all statistics formulas in {EXCEL_OUT}")

if __name__ == '__main__':
    fix_statistics()
