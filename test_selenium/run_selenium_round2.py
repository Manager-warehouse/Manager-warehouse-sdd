# -*- coding: utf-8 -*-
"""
Master Selenium E2E Test Suite Runner for Round 2 System Testing
Drives a real Chrome browser through login + each of the 10 WMS module
landing pages and records genuine reachability results into
docs/test/test_final.xlsx & result_test.md.

Scope note: this is a page-level smoke check (login as the right role,
navigate to the module's landing page, confirm RBAC didn't bounce us to
/login or /forbidden). It does NOT execute the individual business-flow
assertions described by each test case row (e.g. "create receipt", "QC
fail triggers quarantine") -- that would need dedicated per-flow Selenium
scripts and locators that do not exist yet. Every TC under a module is
recorded with the same module-level result.
"""

import os
import sys
import time
import requests
from datetime import datetime
from pathlib import Path

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(Path(__file__).resolve().parent))

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

from config.config import APP_URL, API_URL, HEADLESS, ADMIN_USER, CEO_USER
from pages.wms_pages import LoginPage, ModulePage
from utils.excel_reporter import update_round2_excel, update_round2_markdown
from utils.error_tracer import tracer

SCREENSHOT_DIR = Path(__file__).resolve().parent / "screenshots"

ROLE_CREDENTIALS = {
    "ADMIN": ADMIN_USER,
    "CEO": CEO_USER,
}

# (module code, module name, landing path, role required by that route in
# frontend/src/routes/AppRoutes.jsx). Paths and roles were checked against
# AppRoutes.jsx directly -- several of the original paths here (e.g.
# /outbound/orders, /transfer/requests, /stocktake/list, /pricing/history,
# /returns/list, /reports/ceo) do not exist as routes and always 404'd.
MODULE_SPECS = [
    ("AUTH-001", "Security, Auth & RBAC", "/admin/users", "ADMIN"),
    ("MDM-002", "Master Data Management", "/admin/products", "CEO"),
    ("RCV-003", "Inbound Receipt & QC", "/inbound/receipts", "CEO"),
    ("OUT-004", "Outbound Delivery & POD", "/outbound/delivery-orders", "CEO"),
    ("TRF-005", "Inter-Warehouse Transfer", "/transfers/requests", "CEO"),
    ("STK-006", "Stocktake & Adjustment", "/stocktake", "CEO"),
    ("PRC-007", "Pricing & COGS", "/finance/price-list", "CEO"),
    ("FIN-008", "Finance, Billing & Closing", "/finance/invoices", "CEO"),
    ("RET-009", "Returns, Scrap & Disposal", "/inbound/returns", "CEO"),
    ("RPT-010", "Reports, Dashboard & Alerts", "/reports/ceo-dashboard", "CEO"),
]


def check_app_online():
    """Verify frontend and backend servers are running."""
    try:
        r_fe = requests.get(APP_URL, timeout=3)
        r_be = requests.get(f"{API_URL}/actuator/health", timeout=3)
        return r_fe.status_code in (200, 304, 404), r_be.status_code in (200, 401, 404)
    except Exception:
        return False, False


def build_driver():
    options = Options()
    if HEADLESS:
        options.add_argument("--headless=new")
    options.add_argument("--window-size=1440,900")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)


def login_as(driver, role_name, credentials):
    if not credentials.get("username") or not credentials.get("password"):
        return False, (
            f"No credentials configured for role {role_name} "
            f"(set WMS_{role_name}_EMAIL / WMS_{role_name}_PASSWORD)"
        )
    LoginPage(driver).login(credentials["username"], credentials["password"])
    if "/login" in driver.current_url:
        return False, f"Login rejected for {credentials['username']} (check credentials)"
    return True, "ok"


def run_module_checks(driver):
    """Returns {code: (passed: bool, detail: str)}."""
    module_page = ModulePage(driver)
    results = {}
    session_role = None

    for code, name, path, role in MODULE_SPECS:
        print(f"\n[Selenium E2E] Verifying [{code}] - {name} @ {path} (role: {role}) ...")

        if session_role != role:
            ok, detail = login_as(driver, role, ROLE_CREDENTIALS.get(role, {}))
            session_role = role if ok else None
            if not ok:
                results[code] = (False, detail)
                print(f"  -> SKIPPED: {detail}")
                continue

        passed, current_url, reason = module_page.check_page_loaded(path)
        detail = f"Resolved URL: {current_url}" if passed else reason
        if not passed:
            SCREENSHOT_DIR.mkdir(exist_ok=True)
            shot = SCREENSHOT_DIR / f"{code}.png"
            driver.save_screenshot(str(shot))
            detail += f" (screenshot: {shot.name})"
        results[code] = (passed, detail)
        print(f"  -> {'PASSED' if passed else 'FAILED'}: {detail}")

    return results


def write_reports(module_results):
    import openpyxl
    excel_path = ROOT / "docs" / "test" / "test_final.xlsx"
    wb = openpyxl.load_workbook(str(excel_path))

    results_dict = {}
    total_tcs_processed = 0

    for code, name, path, role in MODULE_SPECS:
        if code not in wb.sheetnames:
            continue
        ws = wb[code]
        module_passed, module_detail = module_results.get(code, (False, "Not executed"))
        status = "Passed" if module_passed else "Failed"
        note = f"Selenium E2E module smoke check ({role} role, {path}): {module_detail}"

        module_tcs = []
        for r in range(11, ws.max_row + 1):
            tc_id = ws.cell(r, 1).value
            tc_desc = ws.cell(r, 2).value
            if not tc_id or not str(tc_id).strip().startswith(code):
                continue
            tc_id_str = str(tc_id).strip()

            module_tcs.append({"id": tc_id_str, "status": status, "note": note})
            tracer.log_test(
                tc_id=tc_id_str,
                module_code=code,
                test_name=str(tc_desc or tc_id_str),
                status=status,
                note=note,
            )

        results_dict[code] = module_tcs
        total_tcs_processed += len(module_tcs)
        print(f"  -> Recorded {len(module_tcs)} TCs for {code} as {status} (module-level smoke result)")

    print("\n[Report Updater] Writing Selenium Round 2 results...")
    p, f = update_round2_excel(results_dict)
    update_round2_markdown(results_dict)
    tracer.generate_report()
    return p, f, total_tcs_processed, excel_path


def run_selenium_round2_suite():
    print("=" * 65)
    print("SELENIUM E2E AUTOMATION SUITE — ROUND 2 SYSTEM TEST")
    print(f"Run Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Target Frontend URL: {APP_URL}")
    print("=" * 65)

    fe_online, be_online = check_app_online()
    print(f"[System Health Check] Frontend Online: {fe_online} | Backend Online: {be_online}")
    if not (fe_online and be_online):
        print("\n[FATAL] App not reachable at the configured URLs. Start the frontend/backend")
        print("        (or set WMS_APP_URL / WMS_API_URL) before running this suite.")
        return

    driver = build_driver()
    try:
        module_results = run_module_checks(driver)
    finally:
        driver.quit()

    p, f, total_tcs_processed, excel_path = write_reports(module_results)

    print("\n" + "=" * 65)
    print("SELENIUM ROUND 2 EXECUTION COMPLETE!")
    print(f"Total TCs tested in Round 2 : {total_tcs_processed}")
    print(f"Round 2 Passed             : {p}")
    print(f"Round 2 Failed             : {f}")
    print(f"Output Excel               : {excel_path}")
    print("=" * 65)


if __name__ == "__main__":
    run_selenium_round2_suite()
