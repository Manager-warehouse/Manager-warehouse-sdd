# -*- coding: utf-8 -*-
"""
Excel and Markdown Report Updater for Selenium Round 2 System Testing
"""

import sys
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent.parent
EXCEL_PATH = ROOT / "docs" / "test" / "test_final.xlsx"
MD_PATH = ROOT / "docs" / "test" / "result_test.md"
RUN_DATE = datetime.now().strftime("%Y-%m-%d")
TESTER_NAME = "Selenium Automation"

STATUS_FILLS = {
    "Passed": PatternFill("solid", fgColor="C6EFCE"),
    "Failed": PatternFill("solid", fgColor="FFC7CE"),
    "Pending": PatternFill("solid", fgColor="FFEB9C"),
    "N/A": PatternFill("solid", fgColor="D9D9D9"),
}

tahoma_regular = Font(name="Tahoma", size=10, bold=False)
tahoma_bold = Font(name="Tahoma", size=10, bold=True)
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def update_round2_excel(results_dict: dict):
    """
    results_dict format:
    {
       "AUTH-001": [
           {"id": "AUTH-001-001", "status": "Passed", "note": "..."},
           ...
       ],
       ...
    }
    """
    if not EXCEL_PATH.exists():
        print(f"Excel file not found: {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(str(EXCEL_PATH))

    total_r2_passed = 0
    total_r2_failed = 0

    for sheet_name, tc_list in results_dict.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Map TC ID to row index
        tc_row_map = {}
        for r in range(11, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val and str(val).strip().startswith(sheet_name):
                tc_row_map[str(val).strip()] = r

        for tc in tc_list:
            tc_id = tc.get("id")
            st = tc.get("status", "Passed")
            note = tc.get("note", "")

            if st == "Passed":
                total_r2_passed += 1
            elif st == "Failed":
                total_r2_failed += 1

            if tc_id in tc_row_map:
                r = tc_row_map[tc_id]
                # Round 2 Status (Col 9 / I)
                c9 = ws.cell(row=r, column=9)
                c9.value = st
                if st in STATUS_FILLS:
                    c9.fill = STATUS_FILLS[st]
                c9.font = tahoma_regular
                c9.alignment = align_center
                c9.border = thin_border

                # Round 2 Test Date (Col 10 / J)
                c10 = ws.cell(row=r, column=10)
                c10.value = RUN_DATE
                c10.font = tahoma_regular
                c10.alignment = align_center
                c10.border = thin_border

                # Round 2 Tester (Col 11 / K)
                c11 = ws.cell(row=r, column=11)
                c11.value = TESTER_NAME
                c11.font = tahoma_regular
                c11.alignment = align_center
                c11.border = thin_border

                # Note (Col 15 / O)
                if note:
                    c15 = ws.cell(row=r, column=15)
                    old_note = str(c15.value or '')
                    c15.value = f"{old_note} | Selenium R2: {note}".strip(" |")
                    c15.font = tahoma_regular
                    c15.alignment = align_left

        # Update Round 2 formula counts in Module sheet header row 7
        ws['B7'] = '=COUNTIF($I$11:$I$5000, "Passed")'
        ws['C7'] = '=COUNTIF($I$11:$I$5000, "Failed")'
        ws['D7'] = '=COUNTIF($I$11:$I$5000, "Pending")'
        ws['E7'] = '=COUNTIF($I$11:$I$5000, "N/A")'

    wb.save(str(EXCEL_PATH))
    print(f"[Excel Report] Successfully updated Round 2 in {EXCEL_PATH}")
    print(f"  -> Round 2 Selenium Passed: {total_r2_passed}, Failed: {total_r2_failed}")
    return total_r2_passed, total_r2_failed


def update_round2_markdown(results_dict: dict):
    if not MD_PATH.exists():
        return

    content = MD_PATH.read_text(encoding="utf-8")
    lines = content.splitlines()

    # Append Selenium Round 2 execution summary section
    summary_section = [
        "",
        "---",
        "",
        f"## 3. KẾT QUẢ KHIỂM THỬ SELENIUM E2E (ROUND 2 SYSTEM TEST)",
        f"**Ngày chạy Selenium:** {RUN_DATE}",
        f"**Công cụ thực thi:** Selenium WebDriver (Python/Pytest)",
        f"**Người thực hiện / Script:** {TESTER_NAME}",
        "",
        "| Module Code | Tên Module | Selenium Tested | Passed | Failed | Status |",
        "|---|---|---|---|---|---|",
    ]

    for sheet_name, tc_list in results_dict.items():
        p = sum(1 for t in tc_list if t.get("status") == "Passed")
        f = sum(1 for t in tc_list if t.get("status") == "Failed")
        tot = len(tc_list)
        st_str = "✅ PASS 100%" if f == 0 else f"⚠️ {f} FAILED"
        summary_section.append(f"| {sheet_name} | {sheet_name} E2E Suite | {tot} | {p} | {f} | {st_str} |")

    new_content = "\n".join(lines + summary_section)
    MD_PATH.write_text(new_content, encoding="utf-8")
    print(f"[Markdown Report] Updated Selenium Round 2 section in {MD_PATH}")
