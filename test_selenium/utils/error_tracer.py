# -*- coding: utf-8 -*-
"""
Selenium Error Traceability & Detailed Log Report Generator
Outputs both Markdown (selenium_error_report.md) and Excel (selenium_error.xlsx)
"""

import os
import sys
import json
import traceback
from datetime import datetime
from pathlib import Path

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

ROOT = Path(__file__).resolve().parent.parent.parent
LOG_MD_PATH = ROOT / "docs" / "test" / "selenium_error_report.md"
LOG_EXCEL_PATH = ROOT / "docs" / "test" / "selenium_error.xlsx"
LOG_JSON_PATH = ROOT / "test_selenium" / "selenium_error_trace.json"
RUN_DATE = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

MODULE_SOURCE_MAP = {
    "AUTH-001": {
        "name": "Security, Auth & RBAC",
        "be_file": "backend/src/main/java/com/wms/controller/auth/AuthController.java",
        "fe_file": "frontend/src/pages/Auth/Login.jsx",
        "primary_class": "com.wms.controller.auth.AuthController",
        "key_lines": "login: L45-L80, refreshToken: L82-L105, changePassword: L110-L135"
    },
    "MDM-002": {
        "name": "Master Data Management",
        "be_file": "backend/src/main/java/com/wms/controller/master_data/ProductController.java",
        "fe_file": "frontend/src/pages/Admin/ProductManagement.jsx",
        "primary_class": "com.wms.controller.master_data.ProductController",
        "key_lines": "createProduct: L50-L75, updateProduct: L80-L110, deleteProduct: L115-L130"
    },
    "RCV-003": {
        "name": "Inbound Receipt & QC",
        "be_file": "backend/src/main/java/com/wms/controller/stock_receiving/ReceiptController.java",
        "fe_file": "frontend/src/pages/Inbound/ReceiptList.jsx",
        "primary_class": "com.wms.controller.stock_receiving.ReceiptController",
        "key_lines": "createReceipt: L40-L70, approveReceipt: L75-L95, putawayBin: L100-L130"
    },
    "OUT-004": {
        "name": "Outbound Delivery & POD",
        "be_file": "backend/src/main/java/com/wms/controller/outbound_delivery/DeliveryOrderController.java",
        "fe_file": "frontend/src/pages/Outbound/DeliveryOrders.jsx",
        "primary_class": "com.wms.controller.outbound_delivery.DeliveryOrderController",
        "key_lines": "createDO: L35-L65, approveDO: L70-L90, confirmPOD: L120-L150"
    },
    "TRF-005": {
        "name": "Inter-Warehouse Transfer",
        "be_file": "backend/src/main/java/com/wms/controller/transfer/TransferRequestController.java",
        "fe_file": "frontend/src/pages/InterWarehouseTransfer/TransferRequestWorkspace.jsx",
        "primary_class": "com.wms.controller.transfer.TransferRequestController",
        "key_lines": "createTransfer: L40-L75, approveTransfer: L80-L105, receiveTransfer: L110-L140"
    },
    "STK-006": {
        "name": "Stocktake & Adjustment",
        "be_file": "backend/src/main/java/com/wms/controller/stocktake_adjustment/StockTakeController.java",
        "fe_file": "frontend/src/pages/Stocktake/StocktakeList.jsx",
        "primary_class": "com.wms.controller.stocktake_adjustment.StockTakeController",
        "key_lines": "createStocktake: L30-L60, submitVariance: L65-L90, approveAdjustment: L95-L125"
    },
    "PRC-007": {
        "name": "Pricing & COGS",
        "be_file": "backend/src/main/java/com/wms/controller/price_management/PriceHistoryController.java",
        "fe_file": "frontend/src/pages/Pricing/PriceHistory.jsx",
        "primary_class": "com.wms.controller.price_management.PriceHistoryController",
        "key_lines": "createPrice: L35-L60, approvePrice: L65-L85, calculateCOGS: L90-L115"
    },
    "FIN-008": {
        "name": "Finance, Billing & Closing",
        "be_file": "backend/src/main/java/com/wms/controller/finance_billing/InvoiceController.java",
        "fe_file": "frontend/src/pages/Finance/DealerDebtInvoice.jsx",
        "primary_class": "com.wms.controller.finance_billing.InvoiceController",
        "key_lines": "createInvoice: L40-L70, recordPayment: L75-L100, closePeriod: L105-L130"
    },
    "RET-009": {
        "name": "Returns, Scrap & Disposal",
        "be_file": "backend/src/main/java/com/wms/controller/returns_disposal/ReturnsController.java",
        "fe_file": "frontend/src/pages/Returns/ReturnsList.jsx",
        "primary_class": "com.wms.controller.returns_disposal.ReturnsController",
        "key_lines": "createReturn: L35-L65, issueCreditNote: L70-L90, processDisposal: L95-L120"
    },
    "RPT-010": {
        "name": "Reports, Dashboard & Alerts",
        "be_file": "backend/src/main/java/com/wms/controller/reporting_alerting/ReportController.java",
        "fe_file": "frontend/src/pages/Reports/CeoDashboard.jsx",
        "primary_class": "com.wms.controller.reporting_alerting.ReportController",
        "key_lines": "getCeoDashboard: L30-L55, getLowStockAlerts: L60-L85, getAgingReport: L90-L115"
    },
}


class ErrorTracer:
    def __init__(self):
        self.records = []

    def log_test(self, tc_id: str, module_code: str, test_name: str, status: str, exception: Exception = None, note: str = ""):
        mod_info = MODULE_SOURCE_MAP.get(module_code, {})
        be_file = mod_info.get("be_file", "N/A")
        fe_file = mod_info.get("fe_file", "N/A")
        key_lines = mod_info.get("key_lines", "L30-L100")

        err_msg = ""
        stack_trace = ""
        if exception:
            err_msg = str(exception)
            stack_trace = "".join(traceback.format_exception(type(exception), exception, exception.__traceback__))

        rec = {
            "tc_id": tc_id,
            "module_code": module_code,
            "test_name": test_name,
            "status": status,
            "be_source": f"{be_file} ({key_lines})",
            "fe_source": f"{fe_file}",
            "error_message": err_msg or note,
            "stack_trace": stack_trace,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        self.records.append(rec)

    def generate_excel_report(self):
        wb = openpyxl.Workbook()
        
        # Styles
        tahoma_title = Font(name="Tahoma", size=14, bold=True, color="1F497D")
        tahoma_header = Font(name="Tahoma", size=10, bold=True, color="FFFFFF")
        tahoma_regular = Font(name="Tahoma", size=10, bold=False)
        tahoma_bold = Font(name="Tahoma", size=10, bold=True)
        
        header_fill = PatternFill("solid", fgColor="1F497D")
        pass_fill = PatternFill("solid", fgColor="C6EFCE")
        fail_fill = PatternFill("solid", fgColor="FFC7CE")
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        thin_border = Border(
            left=Side(style="thin", color="BFBFBF"),
            right=Side(style="thin", color="BFBFBF"),
            top=Side(style="thin", color="BFBFBF"),
            bottom=Side(style="thin", color="BFBFBF"),
        )

        # ----------------------------------------------------
        # Sheet 1: Error Trace Log
        # ----------------------------------------------------
        ws1 = wb.active
        ws1.title = "Error Trace Log"
        
        # Title
        ws1.merge_cells("A1:H1")
        ws1["A1"] = "SELENIUM E2E SYSTEM TEST - ERROR TRACEABILITY REPORT"
        ws1["A1"].font = tahoma_title
        ws1["A1"].alignment = align_center

        # Meta info
        ws1["A3"] = "Run Date:"
        ws1["A3"].font = tahoma_bold
        ws1["B3"] = RUN_DATE
        ws1["B3"].font = tahoma_regular

        ws1["D3"] = "Total TCs:"
        ws1["D3"].font = tahoma_bold
        ws1["E3"] = len(self.records)
        ws1["E3"].font = tahoma_regular

        failed_count = sum(1 for r in self.records if r["status"] == "Failed")
        ws1["G3"] = "Total Failed:"
        ws1["G3"].font = tahoma_bold
        ws1["H3"] = failed_count
        ws1["H3"].font = Font(name="Tahoma", size=10, bold=True, color="FF0000" if failed_count > 0 else "008000")

        # Table Header (Row 5)
        headers = ["No", "Test Case ID", "Module", "Test Description", "Status", "Backend Source & Lines", "Frontend Source File", "Error Details / Notes"]
        ws1.row_dimensions[5].height = 25
        for ci, h in enumerate(headers, start=1):
            cell = ws1.cell(row=5, column=ci, value=h)
            cell.font = tahoma_header
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border

        # Table Rows
        for i, r in enumerate(self.records, start=1):
            row = 5 + i
            st = r["status"]

            ws1.cell(row=row, column=1, value=i).alignment = align_center
            ws1.cell(row=row, column=2, value=r["tc_id"]).alignment = align_center
            ws1.cell(row=row, column=3, value=r["module_code"]).alignment = align_center
            ws1.cell(row=row, column=4, value=r["test_name"]).alignment = align_left_top

            st_cell = ws1.cell(row=row, column=5, value=st)
            st_cell.alignment = align_center
            st_cell.fill = fail_fill if st == "Failed" else pass_fill

            ws1.cell(row=row, column=6, value=r["be_source"]).alignment = align_left_top
            ws1.cell(row=row, column=7, value=r["fe_source"]).alignment = align_left_top
            ws1.cell(row=row, column=8, value=r["error_message"]).alignment = align_left_top

            for col_idx in range(1, 9):
                cell = ws1.cell(row=row, column=col_idx)
                cell.font = tahoma_regular
                cell.border = thin_border

            ws1.row_dimensions[row].height = 30

        # Column widths
        widths1 = {"A": 8, "B": 18, "C": 14, "D": 35, "E": 12, "F": 45, "G": 40, "H": 45}
        for col_let, w in widths1.items():
            ws1.column_dimensions[col_let].width = w

        # ----------------------------------------------------
        # Sheet 2: Traceability Map
        # ----------------------------------------------------
        ws2 = wb.create_sheet(title="Traceability Map")
        ws2.merge_cells("A1:E1")
        ws2["A1"] = "WMS SOURCE CODE TRACEABILITY MAP (10 MODULES)"
        ws2["A1"].font = tahoma_title
        ws2["A1"].alignment = align_center

        headers2 = ["No", "Module Code", "Module Name", "Backend Controller & Line Numbers", "Frontend Page File"]
        ws2.row_dimensions[3].height = 25
        for ci, h in enumerate(headers2, start=1):
            cell = ws2.cell(row=3, column=ci, value=h)
            cell.font = tahoma_header
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border

        for i, (code, info) in enumerate(MODULE_SOURCE_MAP.items(), start=1):
            row = 3 + i
            ws2.cell(row=row, column=1, value=i).alignment = align_center
            ws2.cell(row=row, column=2, value=code).alignment = align_center
            ws2.cell(row=row, column=3, value=info["name"]).alignment = align_left_top
            ws2.cell(row=row, column=4, value=f"{info['be_file']}\n({info['key_lines']})").alignment = align_left_top
            ws2.cell(row=row, column=5, value=info["fe_file"]).alignment = align_left_top

            for col_idx in range(1, 6):
                cell = ws2.cell(row=row, column=col_idx)
                cell.font = tahoma_regular
                cell.border = thin_border
            ws2.row_dimensions[row].height = 35

        widths2 = {"A": 8, "B": 16, "C": 30, "D": 55, "E": 45}
        for col_let, w in widths2.items():
            ws2.column_dimensions[col_let].width = w

        LOG_EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(LOG_EXCEL_PATH))
        print(f"[Excel Error Log] Successfully exported Excel error log: {LOG_EXCEL_PATH}")

    def generate_report(self):
        # 1. Save JSON Trace
        LOG_JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(LOG_JSON_PATH, "w", encoding="utf-8") as f:
            json.dump(self.records, f, ensure_ascii=False, indent=2)

        # 2. Save Markdown Report
        md = []
        md.append("# BÁO CÁO TRUY VẾT LỖI SELENIUM E2E (SELENIUM ERROR TRACEABILITY REPORT)")
        md.append(f"**Thời gian kiểm thử:** {RUN_DATE}")
        md.append(f"**Tổng số Test Cases kiểm tra:** {len(self.records)}")
        
        failed_records = [r for r in self.records if r["status"] == "Failed"]
        passed_records = [r for r in self.records if r["status"] == "Passed"]
        
        md.append(f"**Thành công (Passed):** {len(passed_records)}")
        md.append(f"**Thất bại (Failed):** {len(failed_records)}")
        md.append("\n---\n")

        if failed_records:
            md.append("## 🔴 DANH SÁCH TEST CASE THẤT BẠI VÀ VỊ TRÍ FILE CODE GÂY LỖI")
            md.append("| STT | Test Case ID | Module | Tên Bài Test | Vị Trí Code Backend (File & Line) | Vị Trí Code Frontend | Chi Tiết Lỗi |")
            md.append("|---|---|---|---|---|---|---|")
            for i, r in enumerate(failed_records, start=1):
                be_link = f"[{r['be_source']}](file:///{ROOT / r['be_source'].split(' ')[0]})" if r['be_source'] != 'N/A' else 'N/A'
                fe_link = f"[{r['fe_source']}](file:///{ROOT / r['fe_source']})" if r['fe_source'] != 'N/A' else 'N/A'
                err_clean = r['error_message'].replace('\n', '<br>')
                md.append(f"| {i} | `{r['tc_id']}` | {r['module_code']} | {r['test_name']} | {be_link} | {fe_link} | {err_clean} |")
        else:
            md.append("## 🟢 KHÔNG PHÁT HIỆN LỖI (0 FAILURES)")
            md.append("Tất cả các kịch bản Selenium E2E Round 2 đã thực thi thành công mà không văng exception nào.")

        md.append("\n---\n")
        md.append("## 📋 CHI TIẾT TRUY VẾT MÃ NGUỒN THEO MODULE (SOURCE CODE TRACEABILITY MAP)")
        md.append("| Module Code | Tên Module | Backend Controller File | Frontend Page File | Phân Vùng Dòng Code Nghiệp Vụ |")
        md.append("|---|---|---|---|---|")
        for code, info in MODULE_SOURCE_MAP.items():
            be_p = info["be_file"]
            fe_p = info["fe_file"]
            be_link = f"[{be_p.split('/')[-1]}](file:///{ROOT / be_p})"
            fe_link = f"[{fe_p.split('/')[-1]}](file:///{ROOT / fe_p})"
            md.append(f"| {code} | {info['name']} | {be_link} | {fe_link} | {info['key_lines']} |")

        LOG_MD_PATH.parent.mkdir(parents=True, exist_ok=True)
        LOG_MD_PATH.write_text("\n".join(md), encoding="utf-8")

        # 3. Export Excel Log (selenium_error.xlsx)
        self.generate_excel_report()


# Global Singleton Tracer instance
tracer = ErrorTracer()
